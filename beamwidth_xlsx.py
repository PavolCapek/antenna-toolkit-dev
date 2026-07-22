#!/usr/bin/env python3
"""
Beamwidth extractor with XLSX output and per-frequency antenna exports.

- Accepts one or more CST Farfield Source Files (.ffs).
- Computes: beamwidths at 3/6/10/12 dB, absolute gain (dBi via directivity),
  beam efficiency (eta_beam, eta_beam_percent).
- Writes ONE Excel (.xlsx) workbook:
  • For each input file: a data sheet with the tabular outputs (as before).
  • Two radiation-diagram sheets (phi=0 and phi=90) with θ from -180..180 (1° step)
    and one column per frequency containing **relative gain in dB** for that cut (0 dB = cut max).
  • A global "summary" sheet aggregating per-file averages across all frequency samples
    (computed separately for φ=0° and φ=90° rows), plus frequency range.
- Additionally, generates **.ant files (V3 format)** for each available frequency in each input .ffs:
  • Exactly **720 lines** per file.
  • **Lines 1–360**: azimuth plane (φ≈0°), starting at the front of the beam **0..359°**.
  • **Lines 361–720**: elevation plane (φ≈90°), starting at **+90°** and proceeding around the θ circle
    with a contiguous 1° step sequence: **+90, +89, …, -180, +179, …, +91** (360 total samples).
    (This satisfies the requirement that elevation description starts at +90° and wraps through ±180°.)
  • Values written are **relative gain in dB** (0 dB = cut maximum at that frequency).
  • Files are stored in **radiaiton pattern files/ant_files/**, named `<stem>-<freqGHz>.ant`.

- Generates simplified **LinkCalc .ffs files** for each frequency in each input .ffs:
  - Headerless, space-separated `phi`, `theta`, and absolute total-field directivity in dBi.
  - Files are stored in **radiaiton pattern files/linkCalc/**, named `<stem>-<freqGHz>.ffs`.

- Generates one extensionless **NetSim antenna JSON** per input .ffs:
  - All frequencies are stored as MHz patterns with 361 phi rows by 181 theta columns.
  - Files are stored in **radiaiton pattern files/netsim/** and retain their UUID on reruns.

Usage:
  python3 beamwidth_xlsx.py out.xlsx file1.ffs file2.ffs [--smooth 1 --theta-window 8]

Examples:
  python3 beamwidth_xlsx.py results.xlsx dish_A.ffs dish_B.ffs --smooth 5 --theta-window 5
"""
from __future__ import annotations
import argparse
import csv
import json
import math
import re
import uuid
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import numpy as np
from pipeline.atomic import StageWorkspace
from pipeline.progress import emit_progress
from pipeline.stages import RADIATION_PATTERN_FILES_DIR
from workbook_metadata import (
    build_workbook_manifest,
    stable_input_id,
    unique_sheet_name,
    write_workbook_manifest,
)

Row = Tuple[float, float, complex, complex]  # (phi_deg, theta_deg, Etheta, Ephi)
LinkCalcRow = Tuple[float, float, float]  # (phi_deg, theta_deg, gain_dBi)


class FFSParseError(ValueError):
    pass


# ---------------------------
# Helpers
# ---------------------------

def nearest_phi_index_circular(phis_deg: np.ndarray, target_deg: float) -> int:
    ph = np.asarray(phis_deg, float)
    d = np.abs(ph - target_deg)
    d = np.minimum(d, 360.0 - d)
    return int(np.argmin(d))


def first_crossing_theta(thetas_deg: np.ndarray, gains_rel_dB: np.ndarray, ref_db: float) -> float:
    """Return first θ ≥ 0 where relative gain crosses ≤ ref_db (linear interp in dB domain)."""
    ref = float(ref_db)
    for i in range(1, len(thetas_deg)):
        g0, g1 = gains_rel_dB[i - 1], gains_rel_dB[i]
        if g0 > ref and g1 <= ref:
            t0, t1 = thetas_deg[i - 1], thetas_deg[i]
            frac = (ref - g0) / (g1 - g0) if g1 != g0 else 0.0
            return float(t0 + frac * (t1 - t0))
    return float('nan')


def circular_cell_sizes(angles_rad: np.ndarray) -> np.ndarray:
    """Per‑sample angular extents Δ for a circular axis (e.g., φ in [0, 2π))."""
    a = np.asarray(angles_rad, float)
    order = np.argsort(a)
    a = a[order]
    twopi = 2.0 * math.pi
    a_ext = np.concatenate(([a[-1] - twopi], a, [a[0] + twopi]))
    mids = 0.5 * (a_ext[1:] + a_ext[:-1])
    delta = mids[1:] - mids[:-1]
    inv = np.empty_like(order)
    inv[order] = np.arange(order.size)
    return delta[inv]


def linear_cell_sizes(angles_rad: np.ndarray) -> np.ndarray:
    """Per‑sample angular extents Δ for a linear axis (e.g., θ in [0, π])."""
    a = np.asarray(angles_rad, float)
    order = np.argsort(a)
    a = a[order]
    a_ext = np.empty(a.size + 2, float)
    a_ext[1:-1] = a
    a_ext[0] = a[0] - 0.5 * (a[1] - a[0])
    a_ext[-1] = a[-1] + 0.5 * (a[-1] - a[-2])
    mids = 0.5 * (a_ext[1:] + a_ext[:-1])
    delta = mids[1:] - mids[:-1]
    inv = np.empty_like(order)
    inv[order] = np.arange(order.size)
    return delta[inv]


def numeric_cell(value: float, digits: Optional[int] = None) -> Optional[float]:
    """Return a finite numeric value for spreadsheets, otherwise blank."""
    try:
        out = float(value)
    except (TypeError, ValueError):
        return None
    if not math.isfinite(out):
        return None
    return round(out, digits) if digits is not None else out


# --- beam efficiency helpers (aligned with beam_band.py) ---

_DEFAULT_NULL_MIN_DEPTH_DB = 20.0
_DEFAULT_NULL_MIN_SEPARATION_DEG = 3.0

def _odd_or_next(n: int) -> int:
    return n if n % 2 == 1 else (n + 1)


def _rolling_median(x: np.ndarray, w: int) -> np.ndarray:
    if w <= 1:
        return x
    w = _odd_or_next(int(w))
    pad = w // 2
    xp = np.pad(x, (pad, pad), mode='edge')
    try:
        from numpy.lib.stride_tricks import sliding_window_view
        windows = sliding_window_view(xp, w)
        return np.median(windows, axis=-1)
    except Exception:
        kernel = np.ones(w, dtype=float) / float(w)
        return np.convolve(xp, kernel, mode='valid')


def _find_valid_null_along_theta(
    cut: np.ndarray,
    thetas_deg: np.ndarray,
    peak_idx: int,
    direction: int,
    *,
    null_min_depth_db: float = _DEFAULT_NULL_MIN_DEPTH_DB,
    null_min_separation_deg: float = _DEFAULT_NULL_MIN_SEPARATION_DEG,
) -> int:
    """Find the first real null from peak_idx, ignoring shallow main-beam ripple."""
    n = cut.size
    if n == 0:
        return 0
    if direction not in (-1, 1):
        raise ValueError("direction must be -1 or 1")

    edge_idx = 0 if direction < 0 else n - 1
    peak_power = float(cut[peak_idx])
    if not np.isfinite(peak_power) or peak_power <= 0:
        return edge_idx

    theta_peak = float(thetas_deg[peak_idx])
    min_sep = max(0.0, float(null_min_separation_deg))
    depth_db = abs(float(null_min_depth_db))

    start = peak_idx + direction
    stop = 0 if direction < 0 else n - 1
    if start < 1 or start > n - 2:
        return edge_idx

    first_deep_idx: int | None = None
    for i in range(start, stop, direction):
        if abs(float(thetas_deg[i]) - theta_peak) < min_sep:
            continue
        rel_db = 10.0 * math.log10(max(float(cut[i]), 1e-300) / peak_power)
        if first_deep_idx is None and rel_db <= -depth_db:
            first_deep_idx = i
        if cut[i - 1] >= cut[i] <= cut[i + 1] and rel_db <= -depth_db:
            return i

    return first_deep_idx if first_deep_idx is not None else edge_idx


def _first_null_bounds_along_theta(
    cut: np.ndarray,
    thetas_deg: np.ndarray,
    peak_idx: int,
    *,
    null_min_depth_db: float = _DEFAULT_NULL_MIN_DEPTH_DB,
    null_min_separation_deg: float = _DEFAULT_NULL_MIN_SEPARATION_DEG,
) -> tuple[int, int]:
    """Find first valid null bounds to the left/right of peak_idx."""
    iL = _find_valid_null_along_theta(
        cut,
        thetas_deg,
        peak_idx,
        -1,
        null_min_depth_db=null_min_depth_db,
        null_min_separation_deg=null_min_separation_deg,
    )
    iR = _find_valid_null_along_theta(
        cut,
        thetas_deg,
        peak_idx,
        1,
        null_min_depth_db=null_min_depth_db,
        null_min_separation_deg=null_min_separation_deg,
    )
    return iL, iR


def build_mainlobe_mask(power_lin: np.ndarray, thetas_deg: np.ndarray,
                        smooth_w: int = 1, theta_window_deg: float = 8.0,
                        null_min_depth_db: float = _DEFAULT_NULL_MIN_DEPTH_DB,
                        null_min_separation_deg: float = _DEFAULT_NULL_MIN_SEPARATION_DEG) -> tuple[np.ndarray, list[tuple[int, float, float]]]:
    """Return (mask, bounds) where mask=1 in main-lobe between first valid nulls for each φ cut.
    bounds: list of (ip, theta_L_deg, theta_R_deg) for inspection.
    """
    Nphi, Ntheta = power_lin.shape
    idx_global = int(np.nanargmax(power_lin))
    _, it_global = np.unravel_index(idx_global, power_lin.shape)
    mask = np.zeros_like(power_lin, dtype=float)
    bounds: list[tuple[int, float, float]] = []

    for ip in range(Nphi):
        cut = power_lin[ip, :].astype(float)
        if smooth_w and smooth_w > 1:
            cut = _rolling_median(cut, smooth_w)

        thg = thetas_deg[it_global]
        if theta_window_deg and theta_window_deg > 0:
            within = (np.abs(thetas_deg - thg) <= float(theta_window_deg))
            cand_idx = np.where(within)[0]
            if cand_idx.size == 0:
                cand_idx = np.arange(Ntheta)
        else:
            cand_idx = np.arange(Ntheta)

        local_peak_idx = int(cand_idx[np.argmax(cut[cand_idx])])
        iL, iR = _first_null_bounds_along_theta(
            cut,
            thetas_deg,
            local_peak_idx,
            null_min_depth_db=null_min_depth_db,
            null_min_separation_deg=null_min_separation_deg,
        )
        mask[ip, iL:iR + 1] = 1.0
        bounds.append((ip, float(thetas_deg[iL]), float(thetas_deg[iR])))

    return mask, bounds


# ---------------------------
# Broadband .ffs parsing (CST v3 aware + fallback)
# ---------------------------
_FREQ_RE_FALLBACK = re.compile(r"(?i)freq[^0-9]*([0-9.eE+-]+)\s*([GMk]?Hz)?")


def _unit_to_hz(u: Optional[str]) -> float:
    if not u:
        return 1.0
    u = u.strip().lower()
    if u == 'hz':
        return 1.0
    if u == 'khz':
        return 1e3
    if u == 'mhz':
        return 1e6
    if u == 'ghz':
        return 1e9
    return 1.0


def _read_ffs_broadband_unvalidated(path: Path, freq_regex: Optional[str] = None) -> Dict[float, List[Row]]:
    """Parse a CST .ffs file into {freq_Hz: [(phi, theta, Eθ, Eφ), ...]}."""
    text = Path(path).read_text(errors='ignore')
    lines = text.splitlines()

    try:
        def next_nonblank(j: int) -> int:
            while j < len(lines) and not lines[j].strip():
                j += 1
            return j
        j = next(i for i, L in enumerate(lines) if L.strip() == '// #Frequencies') + 1
        j = next_nonblank(j)
        nfreq = int(float(lines[j].strip())); j += 1
        i = next(i for i, L in enumerate(lines) if 'Radiated/Accepted/Stimulated Power' in L) + 1
        freqs_raw: List[float] = []
        for _ in range(nfreq):
            for _ in range(3):
                i = next_nonblank(i); i += 1
            i = next_nonblank(i)
            try:
                freqs_raw.append(float(lines[i].strip()))
            except ValueError:
                freqs_raw.append(float('nan'))
            i += 1
        block_hdr = '// >> Phi, Theta'
        block_indices = [k for k, L in enumerate(lines) if block_hdr in L]
        nblocks = min(len(block_indices), len(freqs_raw))
        freqs_raw, block_indices = freqs_raw[:nblocks], block_indices[:nblocks]
        by_freq: Dict[float, List[Row]] = {}
        for b, hdr_idx in enumerate(block_indices):
            m = hdr_idx - 1
            while m >= 0 and 'Total #phi samples' not in lines[m]:
                m -= 1
            m += 1
            while m < len(lines) and not lines[m].strip():
                m += 1
            dims = re.findall(r"([0-9]+)", lines[m])
            if len(dims) >= 2:
                nphi, ntheta = int(dims[0]), int(dims[1])
            else:
                nphi = ntheta = None  # type: ignore
            r = hdr_idx + 1
            rows: List[Row] = []
            while r < len(lines) and '// >>' not in lines[r]:
                parts = lines[r].split()
                if len(parts) >= 6:
                    try:
                        phi, theta, etr, eti, epr, epi = map(float, parts[:6])
                        rows.append((phi, theta, complex(etr, eti), complex(epr, epi)))
                    except ValueError:
                        pass
                r += 1
            freq_val = freqs_raw[b]
            by_freq.setdefault(freq_val, []).extend(rows)
        keys = list(by_freq.keys())
        finite = [k for k in keys if math.isfinite(k)]
        if finite and np.nanmedian(finite) < 1e6:
            by_freq = {(f * 1e9 if math.isfinite(f) else f): v for f, v in by_freq.items()}
        return by_freq
    except Exception:
        rx = re.compile(freq_regex) if freq_regex else _FREQ_RE_FALLBACK
        by_freq: Dict[float, List[Row]] = {}
        current_freq = None
        for line in lines:
            m = rx.search(line)
            if m and not re.match(r"^[0-9.eE+\-\s]+$", line):
                val = float(m.group(1)); unit = m.group(2) or ''
                current_freq = val * _unit_to_hz(unit)
                by_freq.setdefault(current_freq, [])
                continue
            parts = line.split()
            if len(parts) >= 6:
                try:
                    phi, theta, etr, eti, epr, epi = map(float, parts[:6])
                except ValueError:
                    continue
                if current_freq is None:
                    continue
                by_freq[current_freq].append((phi, theta, complex(etr, eti), complex(epr, epi)))
        return by_freq


def _validate_ffs_dataset(path: Path, by_freq: Dict[float, List[Row]]) -> None:
    if not by_freq:
        raise FFSParseError(f"{path.name}: no far-field frequency blocks were found")
    for frequency, rows in by_freq.items():
        if not math.isfinite(float(frequency)) or float(frequency) <= 0:
            raise FFSParseError(f"{path.name}: frequency values must be finite and positive")
        if not rows:
            raise FFSParseError(f"{path.name}: frequency {frequency:g} Hz has no field samples")
        phis = np.asarray([row[0] for row in rows], dtype=float)
        thetas = np.asarray([row[1] for row in rows], dtype=float)
        fields = np.asarray(
            [component for row in rows for component in (row[2].real, row[2].imag, row[3].real, row[3].imag)]
        )
        if not np.isfinite(phis).all() or not np.isfinite(thetas).all() or not np.isfinite(fields).all():
            raise FFSParseError(f"{path.name}: frequency {frequency:g} Hz contains non-finite samples")
        nphi = len(np.unique(np.round(phis, 6)))
        ntheta = len(np.unique(np.round(thetas, 6)))
        if nphi < 2 or ntheta < 2:
            raise FFSParseError(f"{path.name}: frequency {frequency:g} Hz needs at least two phi and theta samples")
        grid_points = set(zip(np.round(phis, 6), np.round(thetas, 6)))
        if len(grid_points) != len(rows) or nphi * ntheta != len(grid_points):
            raise FFSParseError(f"{path.name}: frequency {frequency:g} Hz has an incomplete phi/theta grid")


def read_ffs_broadband(path: Path, freq_regex: Optional[str] = None) -> Dict[float, List[Row]]:
    try:
        by_freq = _read_ffs_broadband_unvalidated(path, freq_regex=freq_regex)
    except OSError as exc:
        raise FFSParseError(f"{Path(path).name}: could not read input: {exc}") from exc
    except Exception as exc:
        raise FFSParseError(f"{Path(path).name}: could not parse CST far-field data: {exc}") from exc
    _validate_ffs_dataset(Path(path), by_freq)
    return by_freq


def frequency_mhz_value(frequency_hz: float) -> int | float:
    """Convert Hz to MHz, preserving exact integer-MHz samples as integers."""
    frequency_mhz = frequency_hz / 1e6
    rounded = round(frequency_mhz)
    if math.isclose(frequency_mhz, rounded, rel_tol=0.0, abs_tol=1e-9):
        return int(rounded)
    return float(frequency_mhz)


def interpolate_netsim_gain_grid(
    phis_deg: np.ndarray,
    thetas_deg: np.ndarray,
    gain_dbi: np.ndarray,
) -> np.ndarray:
    """Interpolate a gain grid to NetSim's phi 0..360 by theta 0..180 grid."""
    phis = np.asarray(phis_deg, dtype=float)
    thetas = np.asarray(thetas_deg, dtype=float)
    gain = np.asarray(gain_dbi, dtype=float)
    if gain.shape != (len(phis), len(thetas)):
        raise ValueError("NetSim interpolation requires a phi-by-theta gain grid")

    target_thetas = np.arange(181, dtype=float)
    theta_interpolated = np.vstack(
        [np.interp(target_thetas, thetas, gain_row) for gain_row in gain]
    )

    # Collapse the duplicate 360-degree source row onto 0 degrees and then
    # extend both ends so interpolation across the phi seam stays periodic.
    normalized_phis = np.mod(phis, 360.0)
    order = np.argsort(normalized_phis, kind="stable")
    sorted_phis = normalized_phis[order]
    sorted_gain = theta_interpolated[order, :]
    _rounded_unique, first_indices = np.unique(
        np.round(sorted_phis, 9), return_index=True
    )
    base_phis = sorted_phis[first_indices]
    base_gain = sorted_gain[first_indices, :]
    extended_phis = np.concatenate(
        ([base_phis[-1] - 360.0], base_phis, [base_phis[0] + 360.0])
    )
    extended_gain = np.vstack((base_gain[-1, :], base_gain, base_gain[0, :]))

    target_phis = np.arange(361, dtype=float)
    result = np.empty((361, 181), dtype=float)
    for theta_index in range(181):
        result[:, theta_index] = np.interp(
            target_phis, extended_phis, extended_gain[:, theta_index]
        )
    return result


# ---------------------------
# Core computation for one file (returns table rows + patterns)
# ---------------------------

def compute_for_file(ffs_path: Path, smooth: int, theta_window: float):
    """Return workbook patterns plus LinkCalc and NetSim outputs.

    - rows_out: list[list[object]] for the data table.
    - theta_grid: 1D array of θ (deg) from -180..180 (step=1).
    - patterns_phi0: 2D array (len(theta_grid) x Nfreq) relative dB for φ≈0°.
    - patterns_phi90: 2D array (len(theta_grid) x Nfreq) relative dB for φ≈90°.
    - freq_labels: list of column labels (e.g., '5.500 GHz').
    - linkcalc_rows: mapping of frequency Hz to source-ordered `(phi, theta, gain_dBi)` rows.
    - netsim_patterns: frequency-MHz patterns with 361x181 interpolated gain matrices.
    """
    by_freq = read_ffs_broadband(ffs_path)
    thresh = {3.0: ('theta_3dB_half_deg', 'beamwidth_3dB_2sided_deg'),
              6.0: ('theta_6dB_half_deg', 'beamwidth_6dB_2sided_deg'),
              10.0: ('theta_10dB_half_deg', 'beamwidth_10dB_2sided_deg'),
              12.0: ('theta_12dB_half_deg', 'beamwidth_12dB_2sided_deg')}

    rows_out: list[list[object]] = []
    freq_list = sorted(by_freq.keys())
    freq_labels: list[str] = []
    linkcalc_rows: dict[float, list[LinkCalcRow]] = {}
    netsim_patterns: list[dict[str, object]] = []

    # Build θ grid for patterns: -180..180, 1° step
    theta_grid = np.arange(-180.0, 181.0, 1.0)
    pat0_cols = []
    pat90_cols = []

    for freq_Hz in freq_list:
        rows = by_freq[freq_Hz]
        arr = np.asarray(rows, dtype=object)
        if arr.size == 0:
            raise FFSParseError(f"{ffs_path.name}: frequency {freq_Hz:g} Hz has no usable samples")

        phis = np.asarray(arr[:, 0], float)
        thetas = np.asarray(arr[:, 1], float)
        Etheta = np.asarray(arr[:, 2], complex)
        Ephi = np.asarray(arr[:, 3], complex)

        uphi = np.unique(phis)
        utheta = np.unique(thetas)
        nphi, ntheta = len(uphi), len(utheta)
        if nphi * ntheta != arr.shape[0]:
            uphi = np.unique(np.round(phis, 6))
            utheta = np.unique(np.round(thetas, 6))
            nphi, ntheta = len(uphi), len(utheta)
            if nphi * ntheta != arr.shape[0]:
                raise FFSParseError(f"{ffs_path.name}: frequency {freq_Hz:g} Hz has an incomplete grid")

        order = np.lexsort((thetas, phis))
        Etheta = Etheta[order]
        Ephi = Ephi[order]

        # Always interpret values as FIELD: power proportional to |Eθ|^2 + |Eφ|^2
        P = (np.abs(Etheta) ** 2 + np.abs(Ephi) ** 2).reshape(nphi, ntheta)

        phis_grid = uphi
        thetas_grid = utheta

        Pmax = float(np.nanmax(P))
        if not np.isfinite(Pmax) or Pmax <= 0:
            raise FFSParseError(f"{ffs_path.name}: frequency {freq_Hz:g} Hz has no positive field power")
        Grel_dB = 10.0 * np.log10(np.maximum(P, 1e-300) / Pmax)

        phir = np.radians(phis_grid)
        thetar = np.radians(thetas_grid)
        dphi = circular_cell_sizes(phir)
        dtheta = linear_cell_sizes(thetar)
        weights = np.outer(dphi, np.sin(thetar) * dtheta)
        Prad = float(np.sum(P * weights))
        if not np.isfinite(Prad) or Prad <= 0:
            raise FFSParseError(f"{ffs_path.name}: frequency {freq_Hz:g} Hz has invalid radiated power")
        Gdir = (4.0 * math.pi) * (P / Prad)
        Gabs_dBi = 10.0 * np.log10(np.maximum(Gdir, 1e-300))

        # Match NetSim's native converter: 10*log10((|Etheta|^2 + |Ephi|^2) / 30).
        Gnetsim_dBi = 10.0 * np.log10(np.maximum(P / 30.0, 1e-300))
        netsim_patterns.append(
            {
                "data": interpolate_netsim_gain_grid(
                    phis_grid, thetas_grid, Gnetsim_dBi
                ).tolist(),
                "frequency": frequency_mhz_value(freq_Hz),
            }
        )

        # Gabs_dBi follows the sorted grid. Restore the source row order for
        # LinkCalc so its phi/theta traversal matches the originating block.
        gain_source_order = np.empty(arr.shape[0], dtype=float)
        gain_source_order[order] = Gabs_dBi.reshape(-1)
        linkcalc_rows[freq_Hz] = [
            (float(phi), float(theta), float(gain))
            for phi, theta, gain in zip(phis, thetas, gain_source_order)
        ]

        peak_index = np.unravel_index(int(np.nanargmax(Gabs_dBi)), Gabs_dBi.shape) if np.isfinite(Gabs_dBi).any() else None
        global_max_gain_dBi = float(Gabs_dBi[peak_index]) if peak_index is not None else float('nan')
        if peak_index is not None:
            peak_phi = float(phis_grid[peak_index[0]])
            peak_theta = float(thetas_grid[peak_index[1]])
            back_phi = (peak_phi + 180.0) % 360.0
            back_theta = 180.0 - peak_theta
            back_ip = nearest_phi_index_circular(phis_grid, back_phi)
            back_it = int(np.argmin(np.abs(thetas_grid - back_theta)))
            back_gain_dBi = float(Gabs_dBi[back_ip, back_it])
            front_to_back_dB = global_max_gain_dBi - back_gain_dBi if np.isfinite(back_gain_dBi) else float('nan')
        else:
            front_to_back_dB = float('nan')

        mask, _ = build_mainlobe_mask(P, thetas_grid, smooth_w=smooth, theta_window_deg=theta_window)
        P_main = float(np.sum(P * weights * mask))
        eta_beam = (P_main / Prad) if Prad > 0 else float('nan')
        eta_beam_percent = (eta_beam * 100.0) if np.isfinite(eta_beam) else float('nan')

        # Indices for φ cuts
        ip0 = nearest_phi_index_circular(phis_grid, 0.0)
        ip90 = nearest_phi_index_circular(phis_grid, 90.0)

        # Radiation diagrams: build mirrored θ and gains (relative dB)
        def build_mirrored(thetas_deg: np.ndarray, gcut_dB: np.ndarray):
            th = thetas_deg.astype(float)
            g = gcut_dB.astype(float)
            # Mirror around 0: [-th[::-1] (skip duplicate 0), th]
            th_m = np.concatenate((-th[::-1][:-1], th))
            g_m = np.concatenate((g[::-1][:-1], g))
            return th_m, g_m

        th0, g0 = build_mirrored(thetas_grid, Grel_dB[ip0, :])
        th90, g90 = build_mirrored(thetas_grid, Grel_dB[ip90, :])

        # Interpolate onto the common theta_grid
        g0_i = np.interp(theta_grid, th0, g0)
        g90_i = np.interp(theta_grid, th90, g90)
        pat0_cols.append(g0_i)
        pat90_cols.append(g90_i)

        # Build data table rows for φ cuts
        for phi_deg, ip in [(float(phis_grid[ip0]), ip0), (float(phis_grid[ip90]), ip90)]:
            it0 = int(np.argmin(np.abs(thetas_grid - 0.0)))
            th = thetas_grid[it0:]
            gcut_rel = Grel_dB[ip, it0:]

            theta_vals = {}
            for d in sorted(thresh.keys()):
                theta_vals[d] = first_crossing_theta(th, gcut_rel, -d)

            g0_abs = float(Gabs_dBi[ip, it0]) if np.isfinite(Gabs_dBi[ip, it0]) else float('nan')
            gmax_abs = float(np.nanmax(Gabs_dBi[ip, :])) if np.isfinite(Gabs_dBi).any() else float('nan')
            g0_rel = float(Grel_dB[ip, it0]) if it0 < len(thetas_grid) else float('nan')
            gmax_rel = float(np.nanmax(Grel_dB[ip, :]))

            row = [
                numeric_cell(freq_Hz, 9),
                numeric_cell(freq_Hz / 1e9, 9),
                numeric_cell(phi_deg, 3),
            ]
            for d in sorted(thresh.keys()):
                t = theta_vals[d]
                row += [numeric_cell(t, 6), numeric_cell(t * 2.0, 6)]
            row += [
                numeric_cell(g0_abs, 3),
                numeric_cell(gmax_abs, 3),
                numeric_cell(g0_rel, 3),
                numeric_cell(gmax_rel, 3),
                numeric_cell(eta_beam, 9),
                numeric_cell(eta_beam_percent, 3),
                numeric_cell(global_max_gain_dBi, 6),
                numeric_cell(front_to_back_dB, 6),
            ]
            rows_out.append(row)

        freq_labels.append(f"{freq_Hz/1e9:.3f} GHz")

    # Stack patterns (len(theta_grid) x Nfreq)
    patterns_phi0 = np.column_stack(pat0_cols) if pat0_cols else None
    patterns_phi90 = np.column_stack(pat90_cols) if pat90_cols else None
    if not rows_out or patterns_phi0 is None or patterns_phi90 is None or not freq_labels:
        raise FFSParseError(f"{ffs_path.name}: no usable beam or radiation-pattern results were computed")
    return (
        rows_out,
        theta_grid,
        patterns_phi0,
        patterns_phi90,
        freq_labels,
        linkcalc_rows,
        netsim_patterns,
    )


# ---------------------------
# Summary helpers
# ---------------------------

def _append_summary_rows(summary_accum: list[list[object]], file_label: str, rows: list[list[object]]):
    """Compute per-φ averages over all frequency samples for this file and append to accumulator."""
    if not rows:
        return
    arr = np.array(rows, dtype=object)
    # columns:
    # 0=freq_Hz, 1=freq_GHz, 2=phi_cut_deg, 3..16 metrics
    freq_GHz = arr[:, 1].astype(float)
    phi = arr[:, 2].astype(float)
    metrics = arr[:, 3:].astype(float)  # shape (N, 14)

    def add_group(phi_val: float):
        sel = np.where(np.isfinite(phi) & (np.abs(phi - phi_val) < 1e-6))[0]
        if sel.size == 0:
            return
        g_freq = freq_GHz[sel]
        g_metrics = metrics[sel, :]
        means = np.nanmean(g_metrics, axis=0)
        row = [
            file_label,
            numeric_cell(phi_val, 0),
            int(sel.size),
            numeric_cell(np.nanmin(g_freq), 6),
            numeric_cell(np.nanmax(g_freq), 6),
        ]
        row += [numeric_cell(v, 6) for v in means]
        summary_accum.append(row)

    # φ = 0 and 90 summaries
    add_group(0.0)
    add_group(90.0)


# ---------------------------
# .ant writer helpers
# ---------------------------

def _build_ant_sequences(theta_grid: np.ndarray) -> tuple[list[int], list[int]]:
    """Return (az_idx_order, el_idx_order) over theta_grid indices for .ant V3.

    - Azimuth plane (φ≈0°): 360 samples for angles 0..359°.
      We map each az angle A ∈ [0,359] to θ = wrap(A) in [-180,180], then index in theta_grid.
    - Elevation plane (φ≈90°): 360 samples starting at +90° and stepping by -1°
      around the full circle: +90, +89, ..., -180, +179, ..., +91.
    """
    # Build a mapping from θ (int) to index in theta_grid
    tg_int = np.round(theta_grid).astype(int)
    theta_to_idx = {int(v): int(i) for i, v in enumerate(tg_int)}

    # Azimuth: 0..359 mapped to wrapped θ in [-180, 180]
    az_idx_order: list[int] = []
    for a in range(360):
        tw = ((a + 180) % 360) - 180  # wrap
        idx = theta_to_idx.get(int(tw))
        if idx is None:
            raise ValueError(f"theta_grid missing value for wrapped azimuth angle {tw}°")
        az_idx_order.append(idx)

    # Elevation: start at +90, then +89..-180, then +179..+91  => total 360
    el_angles = list(range(90, -181, -1)) + list(range(179, 90, -1))
    if len(el_angles) != 360:
        raise AssertionError("Elevation sequence must have 360 angles")
    el_idx_order: list[int] = []
    for ang in el_angles:
        idx = theta_to_idx.get(int(ang))
        if idx is None:
            raise ValueError(f"theta_grid missing value for elevation angle {ang}°")
        el_idx_order.append(idx)

    return az_idx_order, el_idx_order


def write_ant_files(ant_dir: Path,
                    stem: str,
                    theta_grid: np.ndarray,
                    patterns_phi0: np.ndarray,
                    patterns_phi90: np.ndarray,
                    freq_labels: list[str]) -> None:
    """Write .ant files (V3, 720 lines) for each frequency column of patterns.

    - One file per frequency: `<stem>-<freqGHz>.ant` placed in `ant_dir`.
    - First 360 lines: azimuth φ≈0° cut, 0..359° order.
    - Next 360 lines: elevation φ≈90° cut, starting at +90° per _build_ant_sequences.
    """
    if patterns_phi0 is None or patterns_phi90 is None or theta_grid is None:
        return

    ant_dir.mkdir(parents=True, exist_ok=True)

    az_idx_order, el_idx_order = _build_ant_sequences(theta_grid)

    nfreq = patterns_phi0.shape[1]
    assert nfreq == patterns_phi90.shape[1] == len(freq_labels)

    # Utility to turn label like "5.500 GHz" into "5.500GHz"
    def label_to_suffix(lbl: str) -> str:
        return lbl.replace(' ', '')

    for j in range(nfreq):
        fname = f"{stem}-{label_to_suffix(freq_labels[j])}.ant"
        out_path = ant_dir / fname
        with out_path.open('w', encoding='utf-8') as f:
            # Azimuth (first 360 lines): angles 0..359 mapped to θ indices
            az_col = patterns_phi0[:, j]
            for idx in az_idx_order:
                f.write(f"{az_col[idx]:.3f}\n")
            # Elevation (next 360 lines): sequence starting at +90°
            el_col = patterns_phi90[:, j]
            for idx in el_idx_order:
                f.write(f"{el_col[idx]:.3f}\n")


def frequency_ghz_token(frequency_hz: float) -> str:
    """Return a GHz filename token with up to one-hertz precision."""
    value = f"{frequency_hz / 1e9:.9f}".rstrip('0').rstrip('.')
    return f"{value}GHz"


def write_linkcalc_files(
    linkcalc_dir: Path,
    stem: str,
    rows_by_frequency: dict[float, list[LinkCalcRow]],
) -> list[Path]:
    """Write one headerless, space-separated phi/theta/gain .ffs per frequency."""
    linkcalc_dir.mkdir(parents=True, exist_ok=True)
    outputs: list[Path] = []
    for frequency_hz in sorted(rows_by_frequency):
        out_path = linkcalc_dir / f"{stem}-{frequency_ghz_token(frequency_hz)}.ffs"
        with out_path.open('w', encoding='utf-8', newline='\n') as handle:
            for phi, theta, gain_dbi in rows_by_frequency[frequency_hz]:
                handle.write(f"{phi:.12g} {theta:.12g} {gain_dbi:.4f}\n")
        outputs.append(out_path)
    return outputs


def retained_netsim_id(existing_paths: list[Path]) -> str:
    """Reuse a valid UUID from current or legacy NetSim files, or create one."""
    for existing_path in existing_paths:
        try:
            payload = json.loads(existing_path.read_text(encoding="utf-8"))
            candidate = str(payload.get("id", "")).strip()
            uuid.UUID(candidate)
            return candidate
        except (OSError, ValueError, TypeError, AttributeError, json.JSONDecodeError):
            continue
    return str(uuid.uuid4())


def write_netsim_file(
    netsim_dir: Path,
    existing_netsim_dirs: list[Path],
    name: str,
    patterns: list[dict[str, object]],
) -> Path:
    """Write one extensionless NetSim antenna JSON while retaining its UUID."""
    netsim_dir.mkdir(parents=True, exist_ok=True)
    out_path = netsim_dir / name
    payload = {
        "id": retained_netsim_id([directory / name for directory in existing_netsim_dirs]),
        "name": name,
        "patterns": patterns,
    }
    out_path.write_text(
        json.dumps(payload, indent="\t", ensure_ascii=False, allow_nan=False),
        encoding="utf-8",
    )
    return out_path


# ---------------------------
# Main
# ---------------------------


def build_input_output_map(paths: list[Path]) -> list[dict[str, str]]:
    used_sheets = {"summary", "_antenna_toolkit"}
    stem_counts: dict[str, int] = {}
    for path in paths:
        key = (path.stem or "Data").lower()
        stem_counts[key] = stem_counts.get(key, 0) + 1
    mappings: list[dict[str, str]] = []
    for path in paths:
        stem = path.stem or "Data"
        ant_stem = stem if stem_counts[stem.lower()] == 1 else f"{stem}-{stable_input_id(path)[:8]}"
        mappings.append(
            {
                "input_id": stable_input_id(path),
                "data": unique_sheet_name(path, used_sheets),
                "phi0": unique_sheet_name(path, used_sheets, suffix="_phi0"),
                "phi90": unique_sheet_name(path, used_sheets, suffix="_phi90"),
                "ant_stem": ant_stem,
            }
        )
    return mappings


def _fail_invalid_inputs(errors: list[str]) -> None:
    if errors:
        detail = "\n".join(f"- {error}" for error in errors)
        raise SystemExit(f"Workbook generation failed; no outputs were published:\n{detail}")

def main() -> int:
    ap = argparse.ArgumentParser(description="Beamwidth/Directivity/Beam-efficiency to XLSX with radiation diagrams, .ant, LinkCalc, and NetSim outputs.")
    ap.add_argument('output', type=Path, help='Output file (.xlsx preferred; if .csv, a per-input CSV is created).')
    ap.add_argument('ffs', nargs='+', type=Path, help='One or more CST Farfield Source Files (.ffs)')
    ap.add_argument('--smooth', type=int, default=1,
                    help='Smoothing window (odd, samples) for main-lobe detection; 1 = no smoothing')
    ap.add_argument('--theta-window', type=float, default=8.0,
                    help='Theta window (deg) around global peak to search for local peaks/nulls')
    args = ap.parse_args()

    input_maps = build_input_output_map(args.ffs)
    validation_errors: list[str] = []
    computed_results: list[tuple] = []
    seen_paths: set[Path] = set()
    progress_total = len(args.ffs) + 1
    for index, fpath in enumerate(args.ffs, start=1):
        emit_progress("beam", index, progress_total, f"Processing {fpath.name}")
        resolved = fpath.resolve()
        if resolved in seen_paths:
            validation_errors.append(f"{fpath.name}: the same FFS input was selected more than once")
            continue
        seen_paths.add(resolved)
        try:
            computed_results.append(compute_for_file(fpath, args.smooth, args.theta_window))
        except FFSParseError as exc:
            validation_errors.append(str(exc))
    _fail_invalid_inputs(validation_errors)

    header = ['freq_Hz', 'freq_GHz', 'phi_cut_deg',
              'theta_3dB_half_deg', 'beamwidth_3dB_2sided_deg',
              'theta_6dB_half_deg', 'beamwidth_6dB_2sided_deg',
              'theta_10dB_half_deg', 'beamwidth_10dB_2sided_deg',
              'theta_12dB_half_deg', 'beamwidth_12dB_2sided_deg',
              'gain_at_theta0_dBi', 'max_gain_dBi',
              'gain_at_theta0_rel_dB', 'max_gain_rel_dB',
              'eta_beam', 'eta_beam_percent',
              'global_max_gain_dBi', 'front_to_back_dB']

    # Header for the summary sheet
    summary_header = (
        ['file', 'phi_cut_deg', 'n_freq', 'freq_min_GHz', 'freq_max_GHz'] +
        ['avg_' + h for h in header[3:]]
    )

    if args.output.suffix.lower() == '.xlsx':
        try:
            from openpyxl import Workbook
        except ImportError:
            raise SystemExit("Please install openpyxl:  pip install openpyxl")

        wb = Workbook()
        # remove default sheet
        default_ws = wb.active
        wb.remove(default_ws)

        # collect summary rows for all files
        summary_rows_all: list[list[object]] = []

        args.output.parent.mkdir(parents=True, exist_ok=True)

        stage = StageWorkspace(args.output.parent, "beam")
        radiation_dir = stage.path(RADIATION_PATTERN_FILES_DIR)
        ant_dir = radiation_dir / "ant_files"
        linkcalc_dir = radiation_dir / "linkCalc"
        netsim_dir = radiation_dir / "netsim"
        existing_netsim_dirs = [
            args.output.parent / RADIATION_PATTERN_FILES_DIR / "netsim",
            args.output.parent / "netsim",
        ]
        expected_ant_files: list[Path] = []
        expected_linkcalc_files: list[Path] = []
        expected_netsim_files: list[Path] = []

        for fpath, mapping, result in zip(args.ffs, input_maps, computed_results):
            rows, theta_grid, pat0, pat90, labels, linkcalc_rows, netsim_patterns = result

            # Data sheet
            title = mapping["data"]
            ws = wb.create_sheet(title=title)
            ws.append(header)
            # Group rows: phi=0 first, then phi=90
            phi0_rows = [r for r in rows if float(r[2]) == 0.0]
            phi90_rows = [r for r in rows if float(r[2]) == 90.0]
            for r in phi0_rows + phi90_rows:
                ws.append(r)

            # Append to summary accumulator
            _append_summary_rows(summary_rows_all, fpath.stem, rows)

            # Radiation diagram sheets (if patterns exist)
            if theta_grid is not None and pat0 is not None:
                tname0 = mapping["phi0"]
                ws0 = wb.create_sheet(title=tname0)
                ws0.append(["theta_deg (deg) — REL dB (0 dB=max)"] + labels)
                for i, th in enumerate(theta_grid):
                    row = [numeric_cell(th, 0)] + [numeric_cell(pat0[i, j], 3) for j in range(pat0.shape[1])]
                    ws0.append(row)

            if theta_grid is not None and pat90 is not None:
                tname90 = mapping["phi90"]
                ws90 = wb.create_sheet(title=tname90)
                ws90.append(["theta_deg (deg) — REL dB (0 dB=max)"] + labels)
                for i, th in enumerate(theta_grid):
                    row = [numeric_cell(th, 0)] + [numeric_cell(pat90[i, j], 3) for j in range(pat90.shape[1])]
                    ws90.append(row)

            # --- Write .ant files for this input (reusing in-memory patterns; no re-parsing) ---
            if theta_grid is not None and pat0 is not None and pat90 is not None and labels:
                expected_ant_files.extend(
                    Path(RADIATION_PATTERN_FILES_DIR)
                    / "ant_files"
                    / f"{mapping['ant_stem']}-{label.replace(' ', '')}.ant"
                    for label in labels
                )
                write_ant_files(
                    ant_dir=ant_dir,
                    stem=mapping["ant_stem"],
                    theta_grid=theta_grid,
                    patterns_phi0=pat0,
                    patterns_phi90=pat90,
                    freq_labels=labels,
                )

            linkcalc_outputs = write_linkcalc_files(
                linkcalc_dir=linkcalc_dir,
                stem=mapping["ant_stem"],
                rows_by_frequency=linkcalc_rows,
            )
            expected_linkcalc_files.extend(
                Path(RADIATION_PATTERN_FILES_DIR) / "linkCalc" / path.name
                for path in linkcalc_outputs
            )
            netsim_output = write_netsim_file(
                netsim_dir=netsim_dir,
                existing_netsim_dirs=existing_netsim_dirs,
                name=mapping["ant_stem"],
                patterns=netsim_patterns,
            )
            expected_netsim_files.append(
                Path(RADIATION_PATTERN_FILES_DIR) / "netsim" / netsim_output.name
            )

        # Finally, write the global summary sheet (once)
        ws_sum = wb.create_sheet(title="summary")
        ws_sum.append(summary_header)
        for r in summary_rows_all:
            ws_sum.append(r)

        write_workbook_manifest(
            wb,
            build_workbook_manifest(
                args.ffs,
                smooth=args.smooth,
                theta_window=args.theta_window,
                sheet_maps=input_maps,
            ),
        )

        emit_progress("beam", progress_total, progress_total, f"Saving {args.output.name}")
        wb.save(stage.path(args.output.name))
        stage.publish(
            [args.output.name, RADIATION_PATTERN_FILES_DIR],
            obsolete=["ant_files", "linkCalc", "netsim"],
            validate=expected_ant_files + expected_linkcalc_files + expected_netsim_files,
        )
        print(f"Wrote XLSX: {args.output} with {len(args.ffs)} file sheets + radiation diagrams + summary.")
        published_radiation_dir = args.output.parent / RADIATION_PATTERN_FILES_DIR
        print(f".ant files written to: {published_radiation_dir / 'ant_files'}")
        print(f"LinkCalc .ffs files written to: {published_radiation_dir / 'linkCalc'}")
        print(f"NetSim antenna files written to: {published_radiation_dir / 'netsim'}")
    else:
        # CSV fallback: data-only, one CSV per input
        base = args.output
        base.parent.mkdir(parents=True, exist_ok=True)
        stage = StageWorkspace(base.parent, "beam")
        radiation_dir = stage.path(RADIATION_PATTERN_FILES_DIR)
        ant_dir = radiation_dir / "ant_files"
        linkcalc_dir = radiation_dir / "linkCalc"
        netsim_dir = radiation_dir / "netsim"
        existing_netsim_dirs = [
            base.parent / RADIATION_PATTERN_FILES_DIR / "netsim",
            base.parent / "netsim",
        ]
        expected_ant_files: list[Path] = []
        expected_linkcalc_files: list[Path] = []
        expected_netsim_files: list[Path] = []
        required_outputs: list[str] = [RADIATION_PATTERN_FILES_DIR]
        for fpath, mapping, result in zip(args.ffs, input_maps, computed_results):
            rows, theta_grid, pat0, pat90, labels, linkcalc_rows, netsim_patterns = result
            out_name = f"{base.stem}-{mapping['ant_stem']}.csv"
            out_csv = stage.path(out_name)
            required_outputs.append(out_name)
            with out_csv.open('w', newline='') as f:
                w = csv.writer(f)
                w.writerow(header)
                # Group rows: phi=0 first, then phi=90
                phi0_rows = [r for r in rows if float(r[2]) == 0.0]
                phi90_rows = [r for r in rows if float(r[2]) == 90.0]
                for r in phi0_rows + phi90_rows:
                    w.writerow(r)
            print(f"Prepared CSV: {base.parent / out_name}")

            # Also generate .ant files even in CSV mode (placed next to base path)
            if theta_grid is not None and pat0 is not None and pat90 is not None and labels:
                expected_ant_files.extend(
                    Path(RADIATION_PATTERN_FILES_DIR)
                    / "ant_files"
                    / f"{mapping['ant_stem']}-{label.replace(' ', '')}.ant"
                    for label in labels
                )
                write_ant_files(
                    ant_dir=ant_dir,
                    stem=mapping["ant_stem"],
                    theta_grid=theta_grid,
                    patterns_phi0=pat0,
                    patterns_phi90=pat90,
                    freq_labels=labels,
                )
            linkcalc_outputs = write_linkcalc_files(
                linkcalc_dir=linkcalc_dir,
                stem=mapping["ant_stem"],
                rows_by_frequency=linkcalc_rows,
            )
            expected_linkcalc_files.extend(
                Path(RADIATION_PATTERN_FILES_DIR) / "linkCalc" / path.name
                for path in linkcalc_outputs
            )
            netsim_output = write_netsim_file(
                netsim_dir=netsim_dir,
                existing_netsim_dirs=existing_netsim_dirs,
                name=mapping["ant_stem"],
                patterns=netsim_patterns,
            )
            expected_netsim_files.append(
                Path(RADIATION_PATTERN_FILES_DIR) / "netsim" / netsim_output.name
            )
        obsolete_outputs = [
            path.name
            for path in base.parent.glob(f"{base.stem}-*.csv")
            if path.name not in required_outputs
        ]
        stage.publish(
            required_outputs,
            obsolete=obsolete_outputs + ["ant_files", "linkCalc", "netsim"],
            validate=expected_ant_files + expected_linkcalc_files + expected_netsim_files,
        )
        emit_progress("beam", progress_total, progress_total, f"Finalizing {base.name}")
        published_radiation_dir = base.parent / RADIATION_PATTERN_FILES_DIR
        print(f".ant files written to: {published_radiation_dir / 'ant_files'}")
        print(f"LinkCalc .ffs files written to: {published_radiation_dir / 'linkCalc'}")
        print(f"NetSim antenna files written to: {published_radiation_dir / 'netsim'}")

    return 0


if __name__ == '__main__':
    raise SystemExit(main())
