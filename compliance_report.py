#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import math
from datetime import datetime, timezone
from pathlib import Path

from compliance.evidence import EvidenceCollector, write_evidence_pdf
from compliance.engine import analyze_files, parse_omitted_angle_range
from compliance.paths import evidence_pdf_path
from compliance.standards import (
    ETSI_EDITION,
    ETSI_SECTOR_EDITION,
    ETSI_SECTOR_SOURCE_URL,
    ETSI_SOURCE_URL,
    FCC_EDITION,
    FCC_SOURCE_URL,
)
from pipeline.atomic import StageWorkspace
from pipeline.progress import emit_progress


def _clean_excel_value(value):
    if isinstance(value, float) and not math.isfinite(value):
        return None
    return value


OVERVIEW_COLUMNS = (
    ("antenna", "Antenna"),
    ("port", "Port"),
    ("polarization", "Polarization"),
    ("family", "Standard Family"),
    ("band", "Band"),
    ("classification", "Class / Category"),
    ("result", "Result"),
    ("frequencies_checked", "Frequencies Checked"),
    ("minimum_ghz", "Minimum GHz"),
    ("maximum_ghz", "Maximum GHz"),
    ("worst_margin_db", "Worst Margin dB"),
    ("explanation", "Explanation"),
)

FREQUENCY_COLUMNS = (
    ("antenna", "Antenna"),
    ("port", "Port"),
    ("frequency_ghz", "Frequency GHz"),
    ("polarization", "Polarization"),
    ("directivity_dbi", "Directivity dBi"),
    ("family", "Standard Family"),
    ("band", "Band"),
    ("classification", "Class / Category"),
    ("result", "Result"),
    ("margin_db", "Margin dB"),
    ("limiting_component", "Limiting Component"),
    ("location", "Location"),
    ("measured", "Measured"),
    ("limit", "Limit"),
    ("unit", "Unit"),
    ("explanation", "Explanation"),
)

FAMILY_ORDER = {
    "ETSI RPE": 0,
    "ETSI Sector RPE": 1,
    "ETSI XPD": 2,
    "FCC Part 101": 3,
}


def _rollup_sort_key(row: dict[str, object]) -> tuple[object, ...]:
    return (
        str(row.get("source_file", "")),
        str(row.get("port_label", "")),
        str(row.get("polarization", "")),
        FAMILY_ORDER.get(str(row.get("family", "")), 99),
        str(row.get("band", "")),
        str(row.get("classification", "")),
    )


def _overview_rows(rows: list[dict[str, object]]) -> list[dict[str, object]]:
    projected = [
        {
            "antenna": row.get("source_file"),
            "port": row.get("port_label"),
            "polarization": row.get("polarization"),
            "family": row.get("family"),
            "band": row.get("band"),
            "classification": row.get("classification"),
            "result": row.get("status"),
            "frequencies_checked": row.get("frequencies_checked"),
            "minimum_ghz": row.get("frequency_min_ghz"),
            "maximum_ghz": row.get("frequency_max_ghz"),
            "worst_margin_db": row.get("worst_margin_db"),
            "explanation": row.get("note", ""),
        }
        for row in sorted(rows, key=_rollup_sort_key)
    ]
    return projected


def _frequency_rows(rows: list[dict[str, object]]) -> list[dict[str, object]]:
    projected = [
        {
            "antenna": row.get("source_file"),
            "port": row.get("port_label"),
            "frequency_ghz": row.get("frequency_ghz"),
            "polarization": row.get("polarization"),
            "directivity_dbi": row.get("max_directivity_dbi"),
            "family": row.get("family"),
            "band": row.get("band"),
            "classification": row.get("classification"),
            "result": row.get("status"),
            "margin_db": row.get("margin_db"),
            "limiting_component": row.get("limiting_component"),
            "location": row.get("location"),
            "measured": row.get("measured_value"),
            "limit": row.get("limit_value"),
            "unit": row.get("unit"),
            "explanation": row.get("note", ""),
        }
        for row in rows
    ]
    return sorted(
        projected,
        key=lambda row: (
            str(row.get("antenna", "")),
            str(row.get("port", "")),
            float(row.get("frequency_ghz", 0.0) or 0.0),
            FAMILY_ORDER.get(str(row.get("family", "")), 99),
            str(row.get("classification", "")),
        ),
    )


def _write_table_sheet(
    workbook,
    title: str,
    rows: list[dict[str, object]],
    columns: tuple[tuple[str, str], ...],
) -> None:
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    worksheet = workbook.create_sheet(title)
    headers = [key for key, _label in columns]
    display_headers = [label for _key, label in columns]
    worksheet.append(display_headers)
    for row in rows:
        worksheet.append([_clean_excel_value(row.get(header)) for header in headers])
    if not rows:
        worksheet.append(["No applicable requirements or results"])
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    worksheet.row_dimensions[1].height = 30
    worksheet.sheet_view.showGridLines = False
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    status_columns = {index + 1 for index, name in enumerate(headers) if name == "result"}
    pass_fill = PatternFill("solid", fgColor="C6EFCE")
    fail_fill = PatternFill("solid", fgColor="FFC7CE")
    neutral_fill = PatternFill("solid", fgColor="E7E6E6")
    row_border = Border(bottom=Side(style="hair", color="D9E2F3"))
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.border = row_border
            cell.alignment = Alignment(vertical="top")
        for index in status_columns:
            cell = row[index - 1]
            normalized = str(cell.value).strip().upper()
            if normalized in {"PASS", "TRUE"}:
                cell.fill = pass_fill
            elif normalized in {"FAIL", "FALSE"}:
                cell.fill = fail_fill
            elif normalized:
                cell.fill = neutral_fill
    for index, header in enumerate(headers, start=1):
        sample = [display_headers[index - 1]] + [str(row.get(header, "")) for row in rows[:200]]
        width = min(48, max(10, max(len(value) for value in sample) + 2))
        worksheet.column_dimensions[worksheet.cell(1, index).column_letter].width = width
        column_cells = worksheet.iter_cols(min_col=index, max_col=index, min_row=2)
        cells = next(column_cells)
        if header == "explanation":
            worksheet.column_dimensions[worksheet.cell(1, index).column_letter].width = 60
            for cell in cells:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                if cell.value:
                    wrapped_lines = max(1, math.ceil(len(str(cell.value)) / 72))
                    worksheet.row_dimensions[cell.row].height = min(120, max(30, 16 * wrapped_lines + 6))
        if header == "frequencies_checked":
            for cell in cells:
                cell.number_format = "0"
        elif header.endswith("_ghz"):
            for cell in cells:
                cell.number_format = "0.000"
        elif header.endswith(("_db", "_dbi")):
            for cell in cells:
                cell.number_format = "0.00"
        elif header in {"measured", "limit"}:
            for cell in cells:
                cell.number_format = "0.00"


def write_workbook(
    output: Path,
    results: dict[str, list[dict[str, object]]],
    *,
    fmin_ghz: float,
    fmax_ghz: float,
    sector_width_deg: float = 0.0,
    sector_center_ghz: float = 0.0,
    evidence_filename: str | None = None,
    generated_at_utc: str | None = None,
) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    workbook = Workbook()
    workbook.remove(workbook.active)
    _write_table_sheet(workbook, "Overview", _overview_rows(results["rollup"]), OVERVIEW_COLUMNS)
    _write_table_sheet(
        workbook,
        "Frequency Results",
        _frequency_rows(results["per_frequency"]),
        FREQUENCY_COLUMNS,
    )

    methodology = workbook.create_sheet("Methodology")
    if fmin_ghz > 0 and fmax_ghz > 0:
        frequency_window = f"{fmin_ghz:g} to {fmax_ghz:g} GHz"
    elif fmin_ghz > 0:
        frequency_window = f"{fmin_ghz:g} GHz and above"
    elif fmax_ghz > 0:
        frequency_window = f"Up to {fmax_ghz:g} GHz"
    else:
        frequency_window = "All input frequencies"
    generated_at = generated_at_utc or datetime.now(timezone.utc).replace(microsecond=0).isoformat()
    evidence_name = evidence_filename or evidence_pdf_path(output).name
    methodology_rows = [
        ("Generated UTC", generated_at),
        ("ETSI rules", ETSI_EDITION),
        ("ETSI source", ETSI_SOURCE_URL),
        ("ETSI sector rules", ETSI_SECTOR_EDITION),
        ("ETSI sector source", ETSI_SECTOR_SOURCE_URL),
        ("FCC rules", FCC_EDITION),
        ("FCC source", FCC_SOURCE_URL),
        ("Frequency window", frequency_window),
        ("Sample coverage", "Every available input frequency sample inside the selected frequency window is listed. Frequencies outside ETSI or FCC bands are retained as Not applicable."),
        ("Gain convention", "Directivity is used wherever ETSI or FCC refers to antenna gain, by explicit project decision."),
        ("Polarization convention", "Ludwig-3 linear co/cross components; H or V comes from the port label/filename, otherwise the main-beam field is used."),
        ("Plane convention", "CST phi=0/180 is azimuth and phi=90/270 is elevation. Opposite sides are evaluated independently."),
        ("ETSI RPE method", "Actual co/cross directivity is compared with every applicable piecewise-linear RPE mask over its published angular domain."),
        (
            "ETSI sector method",
            "Disabled (sector width is 0)."
            if sector_width_deg <= 0
            else (
                f"Linear single-beam sector RPEs with symmetric elevation are evaluated for a declared {sector_width_deg:g} degree sector. "
                "Co-polar and cross-polar results are expressed relative to the strongest co-polar azimuth sample inside the declared sector. "
                + (
                    f"Declared operating-range centre f0 is {sector_center_ghz:g} GHz."
                    if sector_center_ghz > 0
                    else "Automatic f0 uses the midpoint of the selected, bounded compliance frequency window."
                )
            ),
        ),
        ("ETSI XPD method", "Category 1 uses the azimuth 1 dB beamwidth; Category 2 uses the 3D 1 dB contour; Category 3 conservatively includes the 3 degree main-beam region."),
        ("FCC method", "Compliance requires beamwidth in both planes OR directivity, plus every applicable co-polar suppression bin and any cross-polar/XPD requirement."),
        ("Result interpretation", "A rollup result passes only when every evaluated frequency sample for that class or standard passes."),
        ("Margin interpretation", "A positive margin passes, a negative margin fails, and the smallest margin is the worst case."),
        ("Evidence pack", f"See {evidence_name}. Each page shows the minimum-margin frequency for one Overview result."),
        ("Interpretation", "This is a simulation/data pre-compliance assessment. It is not an accredited measurement report or regulatory certification."),
        ("Coverage note", "ETSI range 8/9 class 4 is listed in the overview but no class 4 actual RPE corner-point figure is published in V2.2.1; it is not auto-assigned."),
    ]
    row_numbers: dict[str, int] = {}
    for key, value in methodology_rows:
        methodology.append([key, value])
        row_numbers[key] = methodology.max_row
    label_fill = PatternFill("solid", fgColor="D9EAF7")
    for row in methodology.iter_rows():
        row[0].font = Font(bold=True, color="1F1F1F")
        row[0].fill = label_fill
        row[1].alignment = Alignment(wrap_text=True, vertical="top")
    for key in ("ETSI source", "ETSI sector source", "FCC source"):
        row_number = row_numbers[key]
        methodology.cell(row_number, 2).hyperlink = methodology.cell(row_number, 2).value
        methodology.cell(row_number, 2).style = "Hyperlink"
    evidence_row = row_numbers["Evidence pack"]
    methodology.cell(evidence_row, 2).hyperlink = evidence_name
    methodology.cell(evidence_row, 2).style = "Hyperlink"
    methodology.sheet_view.showGridLines = False
    methodology.column_dimensions["A"].width = 24
    methodology.column_dimensions["B"].width = 120
    methodology.freeze_panes = "A2"
    workbook.save(output)


def _port_labels(raw: str) -> dict[str, str]:
    if not str(raw or "").strip():
        return {}
    payload = json.loads(raw)
    if not isinstance(payload, dict):
        raise ValueError("--port-labels-json must contain a JSON object")
    return {str(key): str(value) for key, value in payload.items()}


def main() -> int:
    parser = argparse.ArgumentParser(description="Check CST far-field patterns against ETSI point-to-point/sector and FCC Part 101 antenna requirements.")
    parser.add_argument("output", type=Path, help="Output compliance workbook (.xlsx)")
    parser.add_argument("ffs", nargs="+", type=Path, help="CST broadband .ffs input files")
    parser.add_argument("--port-labels-json", default="", help="Optional filename-to-port-label JSON object")
    parser.add_argument("--fmin", type=float, default=0.0, help="Minimum frequency in GHz")
    parser.add_argument("--fmax", type=float, default=0.0, help="Maximum frequency in GHz")
    parser.add_argument(
        "--sector-width",
        type=float,
        default=0.0,
        metavar="DEGREES",
        help="Declared ETSI single-beam sector width (2 alpha); 0 disables sector evaluation",
    )
    parser.add_argument(
        "--sector-center",
        type=float,
        default=0.0,
        metavar="GHZ",
        help="Declared sector operating-range centre f0; 0 selects it automatically",
    )
    parser.add_argument(
        "--omit-angle-range",
        type=parse_omitted_angle_range,
        default=None,
        metavar="MIN-MAX",
        help="Inclusive boresight-angle range omitted from ETSI/FCC pattern comparisons",
    )
    args = parser.parse_args()
    if args.fmin > 0 and args.fmax > 0 and args.fmax <= args.fmin:
        parser.error("--fmax must be greater than --fmin")
    if args.sector_width != 0.0 and not 15.0 <= args.sector_width <= 180.0:
        parser.error("--sector-width must be 0 (disabled) or between 15 and 180 degrees")
    if args.sector_center < 0.0:
        parser.error("--sector-center must be 0 (automatic) or greater than 0 GHz")
    if args.sector_width > 0.0 and args.sector_center == 0.0 and not (args.fmin > 0.0 and args.fmax > 0.0):
        parser.error("sector evaluation requires --sector-center, or both --fmin and --fmax for automatic f0")
    if args.output.suffix.lower() != ".xlsx":
        parser.error("output must use the .xlsx extension")

    emit_progress("compliance", 1, 3, "Analyzing ETSI and FCC requirements")
    evidence_collector = EvidenceCollector()
    results = analyze_files(
        args.ffs,
        port_labels=_port_labels(args.port_labels_json),
        fmin_ghz=args.fmin,
        fmax_ghz=args.fmax,
        omitted_angle_range=args.omit_angle_range,
        sector_width_deg=args.sector_width,
        sector_center_ghz=args.sector_center,
        pattern_observer=evidence_collector.observe,
    )
    evidence_cases = evidence_collector.ordered_cases(sorted(results["rollup"], key=_rollup_sort_key))
    generated_at_utc = datetime.now(timezone.utc).replace(microsecond=0).isoformat()
    evidence_output = evidence_pdf_path(args.output)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    with StageWorkspace(args.output.parent, "compliance") as stage:
        staged_output = stage.path(args.output.name)
        staged_evidence = stage.path(evidence_output.name)
        write_workbook(
            staged_output,
            results,
            fmin_ghz=args.fmin,
            fmax_ghz=args.fmax,
            sector_width_deg=args.sector_width,
            sector_center_ghz=args.sector_center,
            evidence_filename=evidence_output.name,
            generated_at_utc=generated_at_utc,
        )
        emit_progress("compliance", 2, 3, f"Creating {evidence_output.name}")
        write_evidence_pdf(
            staged_evidence,
            evidence_cases,
            input_files=args.ffs,
            fmin_ghz=args.fmin,
            fmax_ghz=args.fmax,
            generated_at_utc=generated_at_utc,
            omitted_angle_range=args.omit_angle_range,
        )
        emit_progress("compliance", 3, 3, f"Saving {args.output.name} and {evidence_output.name}")
        stage.publish([args.output.name, evidence_output.name])
    print(f"Wrote compliance workbook: {args.output}")
    print(f"Wrote compliance evidence: {evidence_output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
