from __future__ import annotations

import hashlib
import json
from pathlib import Path
from typing import Any

from pipeline.versions import BEAM_DATA_VERSION, WORKBOOK_MANIFEST_VERSION


WORKBOOK_METADATA_SHEET = "_antenna_toolkit"


def source_fingerprint(path: str | Path) -> dict[str, object]:
    resolved = Path(path).resolve()
    stat = resolved.stat()
    return {
        "path": str(resolved),
        "size": int(stat.st_size),
        "mtime_ns": int(stat.st_mtime_ns),
    }


def stable_input_id(path: str | Path) -> str:
    normalized = str(Path(path).resolve()).replace("\\", "/").lower()
    return hashlib.sha256(normalized.encode("utf-8")).hexdigest()[:10]


def unique_sheet_name(path: Path, used: set[str], *, suffix: str = "") -> str:
    stem = path.stem or "Data"
    candidate = f"{stem[:31 - len(suffix)]}{suffix}"
    if candidate.lower() not in used:
        used.add(candidate.lower())
        return candidate
    identity_suffix = f"-{stable_input_id(path)[:8]}"
    candidate = f"{stem[:31 - len(suffix) - len(identity_suffix)]}{identity_suffix}{suffix}"
    counter = 2
    while candidate.lower() in used:
        numbered = f"-{stable_input_id(path)[:6]}-{counter}"
        candidate = f"{stem[:31 - len(suffix) - len(numbered)]}{numbered}{suffix}"
        counter += 1
    used.add(candidate.lower())
    return candidate


def build_workbook_manifest(
    sources: list[Path],
    *,
    smooth: int,
    theta_window: float,
    sheet_maps: list[dict[str, str]],
) -> dict[str, Any]:
    return {
        "schema_version": WORKBOOK_MANIFEST_VERSION,
        "beam_data_version": BEAM_DATA_VERSION,
        "settings": {"smooth": int(smooth), "theta_window": float(theta_window)},
        "sources": [
            {
                "input_id": stable_input_id(source),
                "fingerprint": source_fingerprint(source),
                "sheets": dict(sheet_map),
            }
            for source, sheet_map in zip(sources, sheet_maps)
        ],
    }


def write_workbook_manifest(workbook, manifest: dict[str, Any]) -> None:
    if WORKBOOK_METADATA_SHEET in workbook.sheetnames:
        del workbook[WORKBOOK_METADATA_SHEET]
    sheet = workbook.create_sheet(WORKBOOK_METADATA_SHEET)
    sheet["A1"] = json.dumps(manifest, sort_keys=True, separators=(",", ":"))
    sheet.sheet_state = "hidden"


def read_workbook_manifest(path: str | Path) -> dict[str, Any] | None:
    from openpyxl import load_workbook

    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            if WORKBOOK_METADATA_SHEET not in workbook.sheetnames:
                return None
            raw = workbook[WORKBOOK_METADATA_SHEET]["A1"].value
        finally:
            workbook.close()
        payload = json.loads(str(raw or ""))
    except Exception:
        return None
    return payload if isinstance(payload, dict) else None


def workbook_source_entry(manifest: dict[str, Any] | None, path: str | Path) -> dict[str, Any] | None:
    if not isinstance(manifest, dict):
        return None
    input_id = stable_input_id(path)
    sources = manifest.get("sources")
    if not isinstance(sources, list):
        return None
    for entry in sources:
        if isinstance(entry, dict) and entry.get("input_id") == input_id:
            return entry
    return None
