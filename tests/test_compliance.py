from __future__ import annotations

import math
import sys
from pathlib import Path

import numpy as np
import pytest

import compliance_report
from compliance.evidence import EvidenceCollector, write_evidence_pdf
from compliance.engine import (
    Pattern,
    _build_per_frequency_results,
    _build_rollup,
    analyze_files,
    analyze_pattern,
    mask_limits,
    parse_omitted_angle_range,
    pattern_from_rows,
)
from compliance.standards import (
    ETSI_EDITION,
    ETSI_SOURCE_URL,
    FCC_EDITION,
    FCC_SOURCE_URL,
    etsi_profiles_for_frequency,
    etsi_sector_profiles,
    fcc_profiles_for_frequency,
)
from compliance_report import write_workbook


def _synthetic_h_rows() -> list[tuple[float, float, complex, complex]]:
    rows: list[tuple[float, float, complex, complex]] = []
    for phi in (0.0, 90.0, 180.0, 270.0, 360.0):
        phi_rad = math.radians(phi)
        for theta in np.arange(0.0, 181.0, 1.0):
            amplitude = math.exp(-((theta / 0.7) ** 2)) + 1e-6
            rows.append(
                (
                    phi,
                    float(theta),
                    complex(amplitude * math.cos(phi_rad), 0.0),
                    complex(-amplitude * math.sin(phi_rad), 0.0),
                )
            )
    return rows


def test_mask_limits_use_last_value_at_vertical_step() -> None:
    angles = np.asarray([70.0, 99.0, 100.0, 101.0, 180.0])
    limits, valid = mask_limits(((70, -2), (100, -7), (100, -10), (180, -10)), angles)

    assert valid.all()
    assert -7 < limits[1] < -6
    assert limits[2] == -10
    assert limits[3] == -10


def test_standard_profiles_are_frequency_specific() -> None:
    assert [profile.class_name for profile in etsi_profiles_for_frequency(4.7)] == ["1", "2", "3", "4"]
    assert [profile.standard for profile in fcc_profiles_for_frequency(6000)] == ["A", "B1", "B2"]
    assert not fcc_profiles_for_frequency(4700)


def test_sector_profiles_apply_declared_width_f0_and_lower_integer_rounding() -> None:
    profiles = etsi_sector_profiles(6.0, 6.0, 90.0)

    assert [profile.class_name for profile in profiles] == ["SS1", "SS2", "SS3"]
    ss2 = profiles[1]
    assert ss2.co_points == ((0.0, 0.0), (50.0, 0.0), (108.0, -20.0), (153.0, -20.0), (180.0, -25.0))
    assert ss2.cross_points == ((0.0, -20.0), (72.0, -20.0), (102.0, -25.0), (159.0, -25.0), (180.0, -25.0))
    assert profiles[2].cross_points[0] == (0.0, -22.0)
    assert profiles[2].co_points[-1] == (180.0, -29.0)


def test_sector_profiles_cover_each_linear_single_beam_frequency_table() -> None:
    assert [profile.class_name for profile in etsi_sector_profiles(2.0, 2.0, 90.0)] == ["SS"]
    assert [profile.class_name for profile in etsi_sector_profiles(30.0, 30.0, 90.0)] == ["SS1", "SS2a", "SS2b", "SS3", "SS4"]
    assert [profile.class_name for profile in etsi_sector_profiles(42.0, 42.0, 90.0)] == ["SS1", "SS2", "SS3"]
    assert not etsi_sector_profiles(15.0, 15.0, 90.0)


def test_omitted_angle_range_parser_accepts_range_single_angle_and_blank() -> None:
    assert parse_omitted_angle_range("179-180") == (179.0, 180.0)
    assert parse_omitted_angle_range("180") == (180.0, 180.0)
    assert parse_omitted_angle_range("") is None
    with pytest.raises(ValueError, match="0 <= minimum"):
        parse_omitted_angle_range("181-181")


def test_omitted_angle_range_removes_backlobe_sample_from_etsi_and_fcc_checks() -> None:
    phis = np.asarray([0.0, 90.0, 180.0, 270.0])
    thetas = np.arange(0.0, 181.0, 1.0)
    co = np.full((len(phis), len(thetas)), -40.0)
    co[:, 0] = 40.0
    co[:, -1] = -8.15
    pattern = Pattern(
        frequency_hz=6e9,
        phis_deg=phis,
        thetas_deg=thetas,
        total_directivity_dbi=co.copy(),
        co_directivity_dbi=co,
        cross_directivity_dbi=np.full_like(co, -80.0),
        polarization="H",
        polarization_basis="test",
    )

    _, strict_etsi, strict_fcc = analyze_pattern(Path("demo_H.ffs"), "H", pattern)
    _, screened_etsi, screened_fcc = analyze_pattern(
        Path("demo_H.ffs"),
        "H",
        pattern,
        omitted_angle_range=(180.0, 180.0),
    )

    strict_class2 = next(row for row in strict_etsi if row["rpe_class"] == "2")
    screened_class2 = next(row for row in screened_etsi if row["rpe_class"] == "2")
    assert strict_class2["status"] == "FAIL"
    assert strict_class2["limiting_angle_deg"] == 180.0
    assert screened_class2["status"] == "PASS"
    assert screened_class2["note"] == ""

    strict_standard_a = next(row for row in strict_fcc if row["standard"] == "A")
    screened_standard_a = next(row for row in screened_fcc if row["standard"] == "A")
    assert strict_standard_a["status"] == "FAIL"
    assert strict_standard_a["limiting_angle_deg"] == 180.0
    assert screened_standard_a["status"] == "PASS"


def test_pattern_uses_ludwig_three_components_and_directivity() -> None:
    pattern = pattern_from_rows(Path("demo_H.ffs"), 6e9, _synthetic_h_rows())

    assert pattern.polarization == "H"
    assert pattern.polarization_basis == "port label or filename"
    assert np.nanmax(pattern.cross_directivity_dbi) < np.nanmax(pattern.co_directivity_dbi) - 100
    assert pattern.max_directivity_dbi > 20


def test_pattern_analysis_reports_etsi_and_fcc_results() -> None:
    pattern = pattern_from_rows(Path("demo_H.ffs"), 6e9, _synthetic_h_rows())
    summary, etsi_rows, fcc_rows = analyze_pattern(Path("demo_H.ffs"), "H", pattern)

    assert summary["etsi_range"] == "1 (3-14 GHz)"
    assert summary["fcc_band"] == "5925-6425 MHz"
    assert len(etsi_rows) == 4
    assert [row["standard"] for row in fcc_rows] == ["A", "B1", "B2"]
    assert all(row["beam_or_gain_pass"] for row in fcc_rows)


def test_file_analysis_filters_every_available_sample_to_compliance_frequency_window(monkeypatch) -> None:
    rows = _synthetic_h_rows()
    monkeypatch.setattr(
        "compliance.engine.read_ffs_broadband",
        lambda _path: {4e9: rows, 6e9: rows, 8e9: rows},
    )

    results = analyze_files(
        [Path("demo_H.ffs")],
        fmin_ghz=5.0,
        fmax_ghz=7.0,
    )

    assert [row["frequency_ghz"] for row in results["summary"]] == [6.0]


def test_file_analysis_adds_sector_classes_for_every_selected_frequency(monkeypatch) -> None:
    rows = _synthetic_h_rows()
    monkeypatch.setattr(
        "compliance.engine.read_ffs_broadband",
        lambda _path: {5e9: rows, 6e9: rows},
    )

    results = analyze_files(
        [Path("demo_H.ffs")],
        fmin_ghz=5.0,
        fmax_ghz=7.0,
        sector_width_deg=90.0,
    )

    assert len(results["sector"]) == 6
    assert {row["sector_class"] for row in results["sector"]} == {"SS1", "SS2", "SS3"}
    assert {row["sector_center_ghz"] for row in results["sector"]} == {6.0}
    sector_frequency_rows = [row for row in results["per_frequency"] if row["family"] == "ETSI Sector RPE"]
    assert len(sector_frequency_rows) == 6
    assert {row["frequency_ghz"] for row in sector_frequency_rows} == {5.0, 6.0}
    assert len([row for row in results["rollup"] if row["family"] == "ETSI Sector RPE"]) == 3


def test_sector_analysis_requires_declared_or_bounded_automatic_center() -> None:
    with pytest.raises(ValueError, match="declared centre frequency"):
        analyze_files([Path("demo_H.ffs")], sector_width_deg=90.0)


def test_failed_sector_classes_explain_each_failed_component(monkeypatch) -> None:
    rows = []
    for phi in (0.0, 90.0, 180.0, 270.0):
        for theta in np.arange(0.0, 181.0, 1.0):
            rows.append((phi, float(theta), 1.0 + 0.0j, 0.5 + 0.0j))
    monkeypatch.setattr("compliance.engine.read_ffs_broadband", lambda _path: {6e9: rows})

    results = analyze_files([Path("demo_H.ffs")], sector_width_deg=90.0, sector_center_ghz=6.0)
    failed = [row for row in results["sector"] if row["status"] == "FAIL"]

    assert failed
    assert all(str(row["note"]).startswith(f"ETSI Sector Class {row['sector_class']} fails because") for row in failed)
    assert all("above the allowed limit" in str(row["note"]) for row in failed)
    assert all("relative to the sector maximum" in str(row["note"]) for row in failed)


def test_each_failed_etsi_class_has_a_plain_language_note() -> None:
    phis = np.asarray([0.0, 90.0, 180.0, 270.0])
    thetas = np.arange(0.0, 181.0, 1.0)
    pattern = Pattern(
        frequency_hz=6e9,
        phis_deg=phis,
        thetas_deg=thetas,
        total_directivity_dbi=np.zeros((len(phis), len(thetas))),
        co_directivity_dbi=np.zeros((len(phis), len(thetas))),
        cross_directivity_dbi=np.full((len(phis), len(thetas)), -40.0),
        polarization="H",
        polarization_basis="test",
    )

    _, etsi_rows, _ = analyze_pattern(Path("demo_H.ffs"), "H", pattern)
    failed = [row for row in etsi_rows if row["status"] == "FAIL"]

    assert failed
    assert all(str(row["note"]).startswith(f"ETSI Class {row['rpe_class']} fails because") for row in failed)
    assert all("above the allowed limit" in str(row["note"]) for row in failed)
    assert all("measured" in str(row["note"]) and "limit" in str(row["note"]) for row in failed)


def test_each_failed_etsi_xpd_category_has_a_plain_language_rollup_note() -> None:
    phis = np.asarray([0.0, 90.0, 180.0, 270.0])
    thetas = np.arange(0.0, 181.0, 1.0)
    pattern = Pattern(
        frequency_hz=6e9,
        phis_deg=phis,
        thetas_deg=thetas,
        total_directivity_dbi=np.zeros((len(phis), len(thetas))),
        co_directivity_dbi=np.zeros((len(phis), len(thetas))),
        cross_directivity_dbi=np.full((len(phis), len(thetas)), -10.0),
        polarization="H",
        polarization_basis="test",
    )

    summary, etsi_rows, fcc_rows = analyze_pattern(Path("demo_H.ffs"), "H", pattern)
    rollup = _build_rollup([summary], etsi_rows, fcc_rows)
    failed_xpd = [row for row in rollup if row["family"] == "ETSI XPD" and row["status"] == "FAIL"]

    assert failed_xpd
    assert all("fails because" in str(row["note"]) for row in failed_xpd)
    assert all("below the required" in str(row["note"]) for row in failed_xpd)


def test_per_frequency_results_cover_every_applicable_class_and_standard() -> None:
    phis = np.asarray([0.0, 90.0, 180.0, 270.0])
    thetas = np.arange(0.0, 181.0, 1.0)
    pattern = Pattern(
        frequency_hz=6e9,
        phis_deg=phis,
        thetas_deg=thetas,
        total_directivity_dbi=np.zeros((len(phis), len(thetas))),
        co_directivity_dbi=np.zeros((len(phis), len(thetas))),
        cross_directivity_dbi=np.full((len(phis), len(thetas)), -10.0),
        polarization="H",
        polarization_basis="test",
    )

    summary, etsi_rows, fcc_rows = analyze_pattern(Path("demo_H.ffs"), "H", pattern)
    rows = _build_per_frequency_results([summary], etsi_rows, fcc_rows)

    assert len([row for row in rows if row["family"] == "ETSI RPE"]) == 4
    assert len([row for row in rows if row["family"] == "ETSI XPD"]) == 3
    assert len([row for row in rows if row["family"] == "FCC Part 101"]) == 3
    assert all(row["frequency_ghz"] == 6.0 for row in rows)
    failed_fcc = [row for row in rows if row["family"] == "FCC Part 101" and row["status"] == "FAIL"]
    assert failed_fcc
    assert all("fails because" in str(row["note"]) for row in failed_fcc)
    assert all(row.get("limiting_component") for row in failed_fcc)


def test_per_frequency_results_keep_samples_outside_all_standard_bands() -> None:
    phis = np.asarray([0.0, 90.0, 180.0, 270.0])
    thetas = np.arange(0.0, 181.0, 1.0)
    pattern = Pattern(
        frequency_hz=500e6,
        phis_deg=phis,
        thetas_deg=thetas,
        total_directivity_dbi=np.zeros((len(phis), len(thetas))),
        co_directivity_dbi=np.zeros((len(phis), len(thetas))),
        cross_directivity_dbi=np.full((len(phis), len(thetas)), -40.0),
        polarization="H",
        polarization_basis="test",
    )

    summary, etsi_rows, fcc_rows = analyze_pattern(Path("demo_H.ffs"), "H", pattern)
    rows = _build_per_frequency_results([summary], etsi_rows, fcc_rows)
    rollup = _build_rollup([summary], etsi_rows, fcc_rows)

    assert {row["family"] for row in rows} == {"ETSI RPE", "ETSI XPD", "FCC Part 101"}
    assert all(row["status"] == "NOT APPLICABLE" for row in rows)
    assert all(row["frequency_ghz"] == 0.5 for row in rows)
    assert {row["family"] for row in rollup} == {"ETSI RPE", "ETSI XPD", "FCC Part 101"}
    assert all(row["status"] == "NOT APPLICABLE" for row in rollup)
    assert all(row["frequencies_checked"] == 1 for row in rollup)


def test_evidence_collector_keeps_not_applicable_standard_families() -> None:
    phis = np.asarray([0.0, 90.0, 180.0, 270.0])
    thetas = np.arange(0.0, 181.0, 1.0)
    pattern = Pattern(
        frequency_hz=500e6,
        phis_deg=phis,
        thetas_deg=thetas,
        total_directivity_dbi=np.zeros((len(phis), len(thetas))),
        co_directivity_dbi=np.zeros((len(phis), len(thetas))),
        cross_directivity_dbi=np.full((len(phis), len(thetas)), -40.0),
        polarization="H",
        polarization_basis="test",
    )
    summary, etsi_rows, fcc_rows = analyze_pattern(Path("demo_H.ffs"), "H", pattern)
    collector = EvidenceCollector()
    collector.observe(Path("demo_H.ffs"), "H", pattern, summary, etsi_rows, [], fcc_rows)
    rollup = _build_rollup([summary], etsi_rows, fcc_rows)

    cases = collector.ordered_cases(rollup)

    assert {case.family for case in cases} == {"ETSI RPE", "ETSI XPD", "FCC Part 101"}
    assert all(case.status == "NOT APPLICABLE" for case in cases)
    assert all(case.margin_db is None for case in cases)


def test_rollup_keeps_fcc_when_etsi_applies_but_no_fcc_band_exists() -> None:
    phis = np.asarray([0.0, 90.0, 180.0, 270.0])
    thetas = np.arange(0.0, 181.0, 1.0)
    pattern = Pattern(
        frequency_hz=5.9e9,
        phis_deg=phis,
        thetas_deg=thetas,
        total_directivity_dbi=np.zeros((len(phis), len(thetas))),
        co_directivity_dbi=np.zeros((len(phis), len(thetas))),
        cross_directivity_dbi=np.full((len(phis), len(thetas)), -40.0),
        polarization="H",
        polarization_basis="test",
    )
    summary, etsi_rows, fcc_rows = analyze_pattern(Path("demo_H.ffs"), "H", pattern)

    rollup = _build_rollup([summary], etsi_rows, fcc_rows)
    fcc_rollup = [row for row in rollup if row["family"] == "FCC Part 101"]

    assert etsi_rows
    assert not fcc_rows
    assert len(fcc_rollup) == 1
    assert fcc_rollup[0]["status"] == "NOT APPLICABLE"
    assert fcc_rollup[0]["classification"] == "No applicable standard"
    assert "was checked" in str(fcc_rollup[0]["note"])


def test_fcc_failure_note_only_mentions_available_qualification_routes() -> None:
    phis = np.asarray([0.0, 90.0, 180.0, 270.0])
    thetas = np.arange(0.0, 181.0, 1.0)
    pattern = Pattern(
        frequency_hz=2e9,
        phis_deg=phis,
        thetas_deg=thetas,
        total_directivity_dbi=np.zeros((len(phis), len(thetas))),
        co_directivity_dbi=np.zeros((len(phis), len(thetas))),
        cross_directivity_dbi=np.full((len(phis), len(thetas)), -20.0),
        polarization="H",
        polarization_basis="test",
    )

    _, _, fcc_rows = analyze_pattern(Path("demo_H.ffs"), "H", pattern)

    assert fcc_rows
    assert all("beamwidth" in str(row["failure_note"]) for row in fcc_rows)
    assert all("neither beamwidth nor directivity" not in str(row["failure_note"]) for row in fcc_rows)
    assert all(row["gain_pass"] is None for row in fcc_rows)
    assert all(row["xpd_pass"] is None for row in fcc_rows)


def test_compliance_workbook_contains_traceable_sheets(tmp_path: Path) -> None:
    output = tmp_path / "compliance.xlsx"
    results = {
        "rollup": [
            {
                "source_file": "demo.ffs",
                "port_label": "H",
                "polarization": "H",
                "family": "FCC Part 101",
                "band": "5925-6425 MHz",
                "classification": "Standard A",
                "status": "PASS",
                "frequencies_checked": 1,
                "frequency_min_ghz": 6.0,
                "frequency_max_ghz": 6.0,
                "worst_margin_db": 1.5,
                "note": "",
            }
        ],
        "per_frequency": [
            {
                "source_file": "demo.ffs",
                "port_label": "H",
                "frequency_ghz": 6.0,
                "polarization": "H",
                "max_directivity_dbi": 35.0,
                "family": "ETSI RPE",
                "band": "1 (3-14 GHz)",
                "classification": "Class 3",
                "status": "FAIL",
                "note": "ETSI Class 3 fails because the pattern exceeds the limit.",
                "limiting_component": "co-polar azimuth",
                "location": "co-polar azimuth at 20.00 degrees from boresight",
                "measured_value": -20.0,
                "limit_value": -25.0,
                "unit": "dBi",
                "margin_db": -5.0,
            }
        ],
        "summary": [{"source_file": "demo.ffs", "frequency_ghz": 6.0, "fcc_best_standard": "A"}],
        "etsi": [
            {
                "source_file": "demo.ffs",
                "rpe_class": "3",
                "status": "FAIL",
                "note": "ETSI Class 3 fails because the co-polar azimuth pattern is 2.00 dB above the allowed limit.",
            }
        ],
        "sector": [
            {
                "source_file": "demo.ffs",
                "standard_family": "ETSI EN 302 326-3 single-beam sector (linear, symmetric elevation)",
                "sector_class": "SS2",
                "status": "PASS",
            }
        ],
        "fcc": [{"source_file": "demo.ffs", "standard": "A", "status": "PASS"}],
    }

    write_workbook(
        output,
        results,
        fmin_ghz=5.9,
        fmax_ghz=6.1,
        evidence_filename="compliance-evidence.pdf",
        generated_at_utc="2026-08-04T10:00:00+00:00",
    )

    from openpyxl import load_workbook

    workbook = load_workbook(output, data_only=True)
    assert workbook.sheetnames == ["Overview", "Frequency Results", "Methodology"]
    assert [cell.value for cell in workbook["Overview"][1]] == [
        "Standard Family",
        "Standard / Norm",
        "Band",
        "Class / Category",
        "Polarization",
        "Result",
        "Antenna",
        "Port",
        "Frequencies Checked",
        "Minimum GHz",
        "Maximum GHz",
        "Worst Margin dB",
        "Explanation",
    ]
    assert [cell.value for cell in workbook["Frequency Results"][1]] == [
        "Standard Family",
        "Standard / Norm",
        "Band",
        "Class / Category",
        "Frequency GHz",
        "Polarization",
        "Result",
        "Antenna",
        "Port",
        "Directivity dBi",
        "Margin dB",
        "Limiting Component",
        "Location",
        "Measured",
        "Limit",
        "Unit",
        "Explanation",
    ]
    assert workbook["Overview"].cell(2, 2).value == FCC_EDITION
    assert workbook["Overview"].cell(2, 2).hyperlink.target == FCC_SOURCE_URL
    assert workbook["Frequency Results"].cell(2, 2).value == ETSI_EDITION
    assert workbook["Frequency Results"].cell(2, 2).hyperlink.target == ETSI_SOURCE_URL
    assert workbook["Frequency Results"].cell(2, 5).number_format == "0.000"
    assert workbook["Frequency Results"].cell(2, 14).number_format == "0.00"
    assert workbook["Frequency Results"].cell(2, 7).fill.fgColor.rgb.endswith("FFC7CE")
    assert "ETSI Class 3 fails because" in workbook["Frequency Results"].cell(2, 17).value
    methodology = dict(workbook["Methodology"].iter_rows(values_only=True))
    assert "Directivity is used" in methodology["Gain convention"]
    assert methodology["Frequency window"] == "5.9 to 6.1 GHz"
    assert "every evaluated frequency" in methodology["Result interpretation"]
    assert "positive margin" in methodology["Margin interpretation"]
    assert "compliance-evidence.pdf" in methodology["Evidence pack"]
    evidence_row = next(
        row for row in workbook["Methodology"].iter_rows() if row[0].value == "Evidence pack"
    )
    assert evidence_row[1].hyperlink.target == "compliance-evidence.pdf"
    assert "single-beam sector" in methodology["ETSI sector rules"].lower() or "302 326-3" in methodology["ETSI sector rules"]


def test_report_rows_group_classes_before_polarizations() -> None:
    rollup_rows = []
    frequency_rows = []
    for polarization in ("H", "V"):
        for class_number in (1, 2):
            common = {
                "source_file": f"demo_{polarization}.ffs",
                "source_path": f"demo_{polarization}.ffs",
                "port_label": polarization,
                "polarization": polarization,
                "family": "ETSI RPE",
                "band": "1 (3-14 GHz)",
                "classification": f"Class {class_number}",
                "status": "PASS",
            }
            rollup_rows.append(
                {
                    **common,
                    "frequencies_checked": 1,
                    "frequency_min_ghz": 6.0,
                    "frequency_max_ghz": 6.0,
                    "worst_margin_db": 1.0,
                    "note": "",
                }
            )
            frequency_rows.append(
                {
                    **common,
                    "frequency_ghz": 6.0,
                    "max_directivity_dbi": 35.0,
                    "margin_db": 1.0,
                    "note": "",
                }
            )

    overview = compliance_report._overview_rows(rollup_rows)
    frequency = compliance_report._frequency_rows(frequency_rows)
    expected = [("Class 1", "H"), ("Class 1", "V"), ("Class 2", "H"), ("Class 2", "V")]

    assert [(row["classification"], row["polarization"]) for row in overview] == expected
    assert [(row["classification"], row["polarization"]) for row in frequency] == expected
    assert all(row["standard"] == ETSI_EDITION for row in overview + frequency)


def test_evidence_collector_selects_minimum_margin_frequency() -> None:
    collector = EvidenceCollector()
    base_pattern = pattern_from_rows(Path("demo_H.ffs"), 5e9, _synthetic_h_rows(), "H")
    for frequency_ghz, margin in ((5.0, 2.0), (6.0, -1.5)):
        pattern = Pattern(
            frequency_hz=frequency_ghz * 1e9,
            phis_deg=base_pattern.phis_deg,
            thetas_deg=base_pattern.thetas_deg,
            total_directivity_dbi=base_pattern.total_directivity_dbi,
            co_directivity_dbi=base_pattern.co_directivity_dbi,
            cross_directivity_dbi=base_pattern.cross_directivity_dbi,
            polarization="H",
            polarization_basis="test",
        )
        summary = {
            "source_file": "demo_H.ffs",
            "source_path": "demo_H.ffs",
            "port_label": "H",
            "frequency_ghz": frequency_ghz,
            "polarization": "H",
            "_etsi_xpd_categories": {},
        }
        etsi = [
            {
                **summary,
                "etsi_range": "1 (3-14 GHz)",
                "rpe_class": "1",
                "status": "PASS" if margin >= 0 else "FAIL",
                "margin_db": margin,
                "limiting_component": "co-polar azimuth",
                "limiting_angle_deg": 10.0,
                "actual_dbi": 18.0,
                "limit_dbi": 20.0,
                "note": "",
            }
        ]
        collector.observe(Path("demo_H.ffs"), "H", pattern, summary, etsi, [], [])

    rollup = [
        {
            "source_file": "demo_H.ffs",
            "source_path": "demo_H.ffs",
            "port_label": "H",
            "polarization": "H",
            "family": "ETSI RPE",
            "band": "1 (3-14 GHz)",
            "classification": "Class 1",
        }
    ]
    selected = collector.ordered_cases(rollup)

    assert len(selected) == 1
    assert selected[0].frequency_ghz == 6.0
    assert selected[0].margin_db == -1.5


def test_evidence_pdf_contains_title_and_one_page_per_selected_family(tmp_path: Path, monkeypatch) -> None:
    rows = _synthetic_h_rows()
    monkeypatch.setattr("compliance.engine.read_ffs_broadband", lambda _path: {6e9: rows})
    collector = EvidenceCollector()
    results = analyze_files(
        [Path("demo_H.ffs")],
        sector_width_deg=90.0,
        sector_center_ghz=6.0,
        pattern_observer=collector.observe,
    )
    all_cases = collector.ordered_cases(results["rollup"])
    selected_by_family = {}
    for case in all_cases:
        selected_by_family.setdefault(case.family, case)
    selected = list(selected_by_family.values())
    output = tmp_path / "compliance-evidence.pdf"

    write_evidence_pdf(
        output,
        selected,
        input_files=[Path("demo_H.ffs")],
        fmin_ghz=5.9,
        fmax_ghz=6.1,
        generated_at_utc="2026-08-04T10:00:00+00:00",
        omitted_angle_range=(180.0, 180.0),
    )

    import fitz

    with fitz.open(output) as document:
        assert document.page_count == len(selected) + 1
        assert "Standards Compliance Evidence" in document[0].get_text()
        page_text = "\n".join(page.get_text() for page in document[1:])
        assert all(family in page_text for family in selected_by_family)


def test_compliance_outputs_are_not_replaced_when_evidence_generation_fails(tmp_path: Path, monkeypatch) -> None:
    output = tmp_path / "demo-compliance.xlsx"
    evidence = tmp_path / "demo-compliance-evidence.pdf"
    output.write_bytes(b"old workbook")
    evidence.write_bytes(b"old evidence")
    source = tmp_path / "demo_H.ffs"
    source.write_text("source", encoding="utf-8")
    results = {
        "rollup": [],
        "per_frequency": [],
        "summary": [],
        "etsi": [],
        "sector": [],
        "fcc": [],
    }

    monkeypatch.setattr(compliance_report, "analyze_files", lambda *_args, **_kwargs: results)
    monkeypatch.setattr(
        compliance_report,
        "write_workbook",
        lambda staged, *_args, **_kwargs: staged.write_bytes(b"new workbook"),
    )

    def fail_evidence(*_args, **_kwargs):
        raise RuntimeError("simulated evidence failure")

    monkeypatch.setattr(compliance_report, "write_evidence_pdf", fail_evidence)
    monkeypatch.setattr(sys, "argv", ["compliance_report.py", str(output), str(source)])

    with pytest.raises(RuntimeError, match="simulated evidence failure"):
        compliance_report.main()

    assert output.read_bytes() == b"old workbook"
    assert evidence.read_bytes() == b"old evidence"
