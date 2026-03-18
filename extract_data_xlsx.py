#!/usr/bin/env python3
from __future__ import annotations

import argparse
import math
import re
from pathlib import Path

import numpy as np
import pandas as pd

from beamwidth_xlsx import (
    build_mainlobe_mask,
    circular_cell_sizes,
    first_crossing_theta,
    linear_cell_sizes,
    nearest_phi_index_circular,
    read_ffs_broadband,
)
from plot_vswr import calc_vswr, pair_to_complex, read_touchstone


def infer_polarization_label(path: Path) -> str:
    stem = path.stem
    tokens = [token.lower() for token in re.split(r"[_\-\s]+", stem) if token]
    aliases = {
        "horizontal": "Horizontal",
        "vertical": "Vertical",
        "rhcp": "RHCP",
        "lhcp": "LHCP",
        "hcp": "HCP",
        "vpol": "Vertical",
        "hpol": "Horizontal",
    }
    for token in reversed(tokens):
        if token in aliases:
            return aliases[token]
        if token == "h":
            return "Horizontal"
        if token == "v":
            return "Vertical"
    return stem


def finite_mean(values: list[float]) -> float | None:
    arr = np.asarray(values, dtype=float)
    arr = arr[np.isfinite(arr)]
    if arr.size == 0:
        return None
    return float(np.mean(arr))


def finite_max(values: list[float]) -> float | None:
    arr = np.asarray(values, dtype=float)
    arr = arr[np.isfinite(arr)]
    if arr.size == 0:
        return None
    return float(np.max(arr))


def maybe_float(value: float | int | None, digits: int = 6) -> float | None:
    if value is None:
        return None
    try:
        out = float(value)
    except (TypeError, ValueError):
        return None
    if not math.isfinite(out):
        return None
    return round(out, digits)


def filter_rows_by_range(rows: list[dict[str, object]], fmin: float | None, fmax: float | None) -> tuple[list[dict[str, object]], float | None, float | None]:
    if not rows:
        return [], None, None

    freqs = np.asarray([float(row["freq_GHz"]) for row in rows], dtype=float)
    mask = np.ones(freqs.shape, dtype=bool)
    if fmin is not None:
        mask &= freqs >= float(fmin)
    if fmax is not None:
        mask &= freqs <= float(fmax)

    filtered = [row for row, keep in zip(rows, mask) if keep]
    if not filtered:
        filtered = rows

    used_freqs = np.asarray([float(row["freq_GHz"]) for row in filtered], dtype=float)
    return filtered, float(np.min(used_freqs)), float(np.max(used_freqs))


def compute_cut_beamwidth(thetas_deg: np.ndarray, gains_rel_db: np.ndarray, threshold_db: float) -> float:
    it0 = int(np.argmin(np.abs(thetas_deg - 0.0)))
    theta_half = first_crossing_theta(thetas_deg[it0:], gains_rel_db[it0:], -abs(float(threshold_db)))
    if not math.isfinite(theta_half):
        return float("nan")
    return float(theta_half * 2.0)


def compute_ffs_rows(ffs_path: Path, smooth: int, theta_window: float) -> list[dict[str, object]]:
    by_freq = read_ffs_broadband(ffs_path)
    if not by_freq:
        return []

    rows_out: list[dict[str, object]] = []
    polarization = infer_polarization_label(ffs_path)

    for freq_hz in sorted(by_freq.keys()):
        raw_rows = by_freq[freq_hz]
        if not raw_rows:
            continue

        arr = np.asarray(raw_rows, dtype=object)
        phis = np.asarray(arr[:, 0], dtype=float)
        thetas = np.asarray(arr[:, 1], dtype=float)
        etheta = np.asarray(arr[:, 2], dtype=complex)
        ephi = np.asarray(arr[:, 3], dtype=complex)

        uphi = np.unique(np.round(phis, 6))
        utheta = np.unique(np.round(thetas, 6))
        nphi, ntheta = len(uphi), len(utheta)
        if nphi * ntheta != arr.shape[0]:
            continue

        order = np.lexsort((thetas, phis))
        etheta = etheta[order]
        ephi = ephi[order]

        power = (np.abs(etheta) ** 2 + np.abs(ephi) ** 2).reshape(nphi, ntheta)
        pmax = float(np.nanmax(power))
        if not math.isfinite(pmax) or pmax <= 0:
            continue

        grel_db = 10.0 * np.log10(np.maximum(power, 1e-300) / pmax)
        phir = np.radians(uphi)
        thetar = np.radians(utheta)
        dphi = circular_cell_sizes(phir)
        dtheta = linear_cell_sizes(thetar)
        weights = np.outer(dphi, np.sin(thetar) * dtheta)
        prad = float(np.sum(power * weights))
        if not math.isfinite(prad) or prad <= 0:
            continue

        gdir = (4.0 * math.pi) * (power / prad)
        gabs_dbi = 10.0 * np.log10(np.maximum(gdir, 1e-300))
        mask, _ = build_mainlobe_mask(power, utheta, smooth_w=smooth, theta_window_deg=theta_window)
        p_main = float(np.sum(power * weights * mask))
        beam_eff = p_main / prad if prad > 0 else float("nan")
        beam_eff_percent = beam_eff * 100.0 if math.isfinite(beam_eff) else float("nan")

        ip_az = nearest_phi_index_circular(uphi, 0.0)
        ip_el = nearest_phi_index_circular(uphi, 90.0)
        az_bw_3 = compute_cut_beamwidth(utheta, grel_db[ip_az, :], 3.0)
        az_bw_6 = compute_cut_beamwidth(utheta, grel_db[ip_az, :], 6.0)
        el_bw_3 = compute_cut_beamwidth(utheta, grel_db[ip_el, :], 3.0)
        el_bw_6 = compute_cut_beamwidth(utheta, grel_db[ip_el, :], 6.0)

        peak_index = np.unravel_index(int(np.nanargmax(gabs_dbi)), gabs_dbi.shape)
        peak_gain = float(gabs_dbi[peak_index])
        peak_phi = float(uphi[peak_index[0]])
        peak_theta = float(utheta[peak_index[1]])
        back_phi = (peak_phi + 180.0) % 360.0
        back_theta = 180.0 - peak_theta
        back_ip = nearest_phi_index_circular(uphi, back_phi)
        back_it = int(np.argmin(np.abs(utheta - back_theta)))
        back_gain = float(gabs_dbi[back_ip, back_it])
        front_to_back_db = peak_gain - back_gain if math.isfinite(back_gain) else float("nan")

        rows_out.append(
            {
                "source_file": ffs_path.name,
                "source_stem": ffs_path.stem,
                "polarization": polarization,
                "freq_Hz": float(freq_hz),
                "freq_GHz": float(freq_hz) / 1e9,
                "max_gain_dBi": peak_gain,
                "avg_gain_dBi": peak_gain,
                "azimuth_bw_3dB_deg": az_bw_3,
                "azimuth_bw_6dB_deg": az_bw_6,
                "elevation_bw_3dB_deg": el_bw_3,
                "elevation_bw_6dB_deg": el_bw_6,
                "beam_efficiency": beam_eff,
                "beam_efficiency_percent": beam_eff_percent,
                "front_to_back_dB": front_to_back_db,
            }
        )

    return rows_out


def beam_workbook_is_fresh(beam_workbook: Path, ffs_paths: list[Path]) -> bool:
    if not beam_workbook.exists():
        return False
    try:
        beam_mtime = beam_workbook.stat().st_mtime
    except OSError:
        return False
    for ffs_path in ffs_paths:
        try:
            if ffs_path.stat().st_mtime > beam_mtime:
                return False
        except OSError:
            return False
    return True


def compute_ffs_rows_from_beam_workbook(beam_workbook: Path, ffs_path: Path) -> list[dict[str, object]]:
    if not beam_workbook.exists():
        return []

    sheet_name = ffs_path.stem[:31]
    try:
        df = pd.read_excel(beam_workbook, sheet_name=sheet_name)
    except Exception:
        return []

    required = {
        "freq_GHz",
        "phi_cut_deg",
        "beamwidth_3dB_2sided_deg",
        "beamwidth_6dB_2sided_deg",
        "eta_beam",
        "eta_beam_percent",
        "global_max_gain_dBi",
        "front_to_back_dB",
    }
    if not required.issubset(df.columns):
        return []

    polarization = infer_polarization_label(ffs_path)
    az = df[df["phi_cut_deg"] == 0].copy()
    el = df[df["phi_cut_deg"] == 90].copy()
    if az.empty and el.empty:
        return []

    az = az[["freq_GHz", "beamwidth_3dB_2sided_deg", "beamwidth_6dB_2sided_deg"]].rename(
        columns={
            "beamwidth_3dB_2sided_deg": "azimuth_bw_3dB_deg",
            "beamwidth_6dB_2sided_deg": "azimuth_bw_6dB_deg",
        }
    )
    el = el[["freq_GHz", "beamwidth_3dB_2sided_deg", "beamwidth_6dB_2sided_deg"]].rename(
        columns={
            "beamwidth_3dB_2sided_deg": "elevation_bw_3dB_deg",
            "beamwidth_6dB_2sided_deg": "elevation_bw_6dB_deg",
        }
    )
    shared = (
        df[["freq_GHz", "eta_beam", "eta_beam_percent", "global_max_gain_dBi", "front_to_back_dB"]]
        .drop_duplicates(subset=["freq_GHz"])
        .rename(columns={"global_max_gain_dBi": "max_gain_dBi", "eta_beam": "beam_efficiency", "eta_beam_percent": "beam_efficiency_percent"})
    )
    merged = shared.merge(az, on="freq_GHz", how="left").merge(el, on="freq_GHz", how="left")
    merged = merged.sort_values("freq_GHz")

    rows_out: list[dict[str, object]] = []
    for row in merged.to_dict(orient="records"):
        rows_out.append(
            {
                "source_file": ffs_path.name,
                "source_stem": ffs_path.stem,
                "polarization": polarization,
                "freq_Hz": float(row["freq_GHz"]) * 1e9,
                "freq_GHz": float(row["freq_GHz"]),
                "max_gain_dBi": float(row["max_gain_dBi"]) if pd.notna(row["max_gain_dBi"]) else float("nan"),
                "avg_gain_dBi": float(row["max_gain_dBi"]) if pd.notna(row["max_gain_dBi"]) else float("nan"),
                "azimuth_bw_3dB_deg": float(row["azimuth_bw_3dB_deg"]) if pd.notna(row.get("azimuth_bw_3dB_deg")) else float("nan"),
                "azimuth_bw_6dB_deg": float(row["azimuth_bw_6dB_deg"]) if pd.notna(row.get("azimuth_bw_6dB_deg")) else float("nan"),
                "elevation_bw_3dB_deg": float(row["elevation_bw_3dB_deg"]) if pd.notna(row.get("elevation_bw_3dB_deg")) else float("nan"),
                "elevation_bw_6dB_deg": float(row["elevation_bw_6dB_deg"]) if pd.notna(row.get("elevation_bw_6dB_deg")) else float("nan"),
                "beam_efficiency": float(row["beam_efficiency"]) if pd.notna(row["beam_efficiency"]) else float("nan"),
                "beam_efficiency_percent": float(row["beam_efficiency_percent"]) if pd.notna(row["beam_efficiency_percent"]) else float("nan"),
                "front_to_back_dB": float(row["front_to_back_dB"]) if pd.notna(row["front_to_back_dB"]) else float("nan"),
            }
        )
    return rows_out


def summarize_ffs_rows(rows: list[dict[str, object]]) -> dict[str, object] | None:
    if not rows:
        return None
    return {
        "source_file": rows[0]["source_file"],
        "polarization": rows[0]["polarization"],
        "points_used": len(rows),
        "freq_min_GHz": maybe_float(float(rows[0]["freq_GHz"]) if len(rows) == 1 else min(float(row["freq_GHz"]) for row in rows), 6),
        "freq_max_GHz": maybe_float(float(rows[0]["freq_GHz"]) if len(rows) == 1 else max(float(row["freq_GHz"]) for row in rows), 6),
        "max_gain_dBi_in_range": maybe_float(finite_max([float(row["max_gain_dBi"]) for row in rows]), 4),
        "avg_gain_dBi_in_range": maybe_float(finite_mean([float(row["avg_gain_dBi"]) for row in rows]), 4),
        "avg_azimuth_bw_3dB_deg": maybe_float(finite_mean([float(row["azimuth_bw_3dB_deg"]) for row in rows]), 4),
        "avg_azimuth_bw_6dB_deg": maybe_float(finite_mean([float(row["azimuth_bw_6dB_deg"]) for row in rows]), 4),
        "avg_elevation_bw_3dB_deg": maybe_float(finite_mean([float(row["elevation_bw_3dB_deg"]) for row in rows]), 4),
        "avg_elevation_bw_6dB_deg": maybe_float(finite_mean([float(row["elevation_bw_6dB_deg"]) for row in rows]), 4),
        "avg_beam_efficiency_percent": maybe_float(finite_mean([float(row["beam_efficiency_percent"]) for row in rows]), 4),
        "avg_front_to_back_dB": maybe_float(finite_mean([float(row["front_to_back_dB"]) for row in rows]), 4),
    }


def compute_touchstone_rows(path: Path) -> list[dict[str, object]]:
    freqs_hz, data, fmt, z0, nports = read_touchstone(str(path))
    if freqs_hz.size == 0 or data.size == 0:
        return []

    port_map = [(1, 0)]
    if nports == 2:
        port_map = [(1, 0), (2, 6)]

    rows_out: list[dict[str, object]] = []
    freqs_ghz = freqs_hz / 1e9

    for port, start_idx in port_map:
        gamma = np.asarray(
            [pair_to_complex(float(row[start_idx]), float(row[start_idx + 1]), fmt) for row in data],
            dtype=complex,
        )
        vswr = calc_vswr(gamma)
        denom = 1.0 - gamma
        safe = np.where(np.abs(denom) < 1e-12, np.nan + 0j, denom)
        impedance = z0 * (1.0 + gamma) / safe
        for freq_ghz, gamma_value, vswr_value, z_value in zip(freqs_ghz, gamma, vswr, impedance):
            rows_out.append(
                {
                    "touchstone_file": path.name,
                    "port": f"Port {port}",
                    "freq_GHz": float(freq_ghz),
                    "vswr": float(vswr_value),
                    "impedance_real_ohm": float(np.real(z_value)) if np.isfinite(np.real(z_value)) else float("nan"),
                    "impedance_imag_ohm": float(np.imag(z_value)) if np.isfinite(np.imag(z_value)) else float("nan"),
                    "impedance_magnitude_ohm": float(np.abs(z_value)) if np.isfinite(np.abs(z_value)) else float("nan"),
                    "gamma_real": float(np.real(gamma_value)),
                    "gamma_imag": float(np.imag(gamma_value)),
                    "reference_impedance_ohm": float(z0),
                }
            )

    return rows_out


def summarize_touchstone_rows(rows: list[dict[str, object]]) -> dict[str, object] | None:
    if not rows:
        return None
    return {
        "touchstone_file": rows[0]["touchstone_file"],
        "port": rows[0]["port"],
        "points_used": len(rows),
        "freq_min_GHz": maybe_float(min(float(row["freq_GHz"]) for row in rows), 6),
        "freq_max_GHz": maybe_float(max(float(row["freq_GHz"]) for row in rows), 6),
        "max_vswr_in_range": maybe_float(finite_max([float(row["vswr"]) for row in rows]), 4),
        "avg_vswr_in_range": maybe_float(finite_mean([float(row["vswr"]) for row in rows]), 4),
        "avg_impedance_real_ohm": maybe_float(finite_mean([float(row["impedance_real_ohm"]) for row in rows]), 4),
        "avg_impedance_imag_ohm": maybe_float(finite_mean([float(row["impedance_imag_ohm"]) for row in rows]), 4),
        "avg_impedance_magnitude_ohm": maybe_float(finite_mean([float(row["impedance_magnitude_ohm"]) for row in rows]), 4),
        "reference_impedance_ohm": maybe_float(float(rows[0]["reference_impedance_ohm"]), 4),
    }


def write_sheet(ws, header: list[str], rows: list[list[object]]) -> None:
    ws.append(header)
    for row in rows:
        ws.append(row)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def autosize_columns(ws) -> None:
    for column_cells in ws.columns:
        values = ["" if cell.value is None else str(cell.value) for cell in column_cells]
        width = min(max(len(value) for value in values) + 2, 40) if values else 12
        ws.column_dimensions[column_cells[0].column_letter].width = width


def build_workbook(
    output: Path,
    overview_rows: list[list[object]],
    ffs_summaries: list[dict[str, object]],
    ffs_details: list[dict[str, object]],
    ts_summaries: list[dict[str, object]],
    ts_details: list[dict[str, object]],
) -> None:
    from openpyxl import Workbook

    wb = Workbook()
    default_ws = wb.active
    wb.remove(default_ws)

    ws_overview = wb.create_sheet("overview")
    write_sheet(
        ws_overview,
        ["setting", "value"],
        overview_rows,
    )

    if ffs_summaries:
        ws = wb.create_sheet("ffs_summary")
        write_sheet(
            ws,
            [
                "source_file",
                "polarization",
                "points_used",
                "freq_min_GHz",
                "freq_max_GHz",
                "max_gain_dBi_in_range",
                "avg_gain_dBi_in_range",
                "avg_azimuth_bw_3dB_deg",
                "avg_azimuth_bw_6dB_deg",
                "avg_elevation_bw_3dB_deg",
                "avg_elevation_bw_6dB_deg",
                "avg_beam_efficiency_percent",
                "avg_front_to_back_dB",
            ],
            [
                [
                    row["source_file"],
                    row["polarization"],
                    row["points_used"],
                    row["freq_min_GHz"],
                    row["freq_max_GHz"],
                    row["max_gain_dBi_in_range"],
                    row["avg_gain_dBi_in_range"],
                    row["avg_azimuth_bw_3dB_deg"],
                    row["avg_azimuth_bw_6dB_deg"],
                    row["avg_elevation_bw_3dB_deg"],
                    row["avg_elevation_bw_6dB_deg"],
                    row["avg_beam_efficiency_percent"],
                    row["avg_front_to_back_dB"],
                ]
                for row in ffs_summaries
            ],
        )

    if ffs_details:
        ws = wb.create_sheet("ffs_detail")
        write_sheet(
            ws,
            [
                "source_file",
                "polarization",
                "freq_GHz",
                "max_gain_dBi",
                "azimuth_bw_3dB_deg",
                "azimuth_bw_6dB_deg",
                "elevation_bw_3dB_deg",
                "elevation_bw_6dB_deg",
                "beam_efficiency",
                "beam_efficiency_percent",
                "front_to_back_dB",
            ],
            [
                [
                    row["source_file"],
                    row["polarization"],
                    maybe_float(float(row["freq_GHz"]), 6),
                    maybe_float(float(row["max_gain_dBi"]), 4),
                    maybe_float(float(row["azimuth_bw_3dB_deg"]), 4),
                    maybe_float(float(row["azimuth_bw_6dB_deg"]), 4),
                    maybe_float(float(row["elevation_bw_3dB_deg"]), 4),
                    maybe_float(float(row["elevation_bw_6dB_deg"]), 4),
                    maybe_float(float(row["beam_efficiency"]), 6),
                    maybe_float(float(row["beam_efficiency_percent"]), 4),
                    maybe_float(float(row["front_to_back_dB"]), 4),
                ]
                for row in ffs_details
            ],
        )

    if ts_summaries:
        ws = wb.create_sheet("touchstone_summary")
        write_sheet(
            ws,
            [
                "touchstone_file",
                "port",
                "points_used",
                "freq_min_GHz",
                "freq_max_GHz",
                "max_vswr_in_range",
                "avg_vswr_in_range",
                "avg_impedance_real_ohm",
                "avg_impedance_imag_ohm",
                "avg_impedance_magnitude_ohm",
                "reference_impedance_ohm",
            ],
            [
                [
                    row["touchstone_file"],
                    row["port"],
                    row["points_used"],
                    row["freq_min_GHz"],
                    row["freq_max_GHz"],
                    row["max_vswr_in_range"],
                    row["avg_vswr_in_range"],
                    row["avg_impedance_real_ohm"],
                    row["avg_impedance_imag_ohm"],
                    row["avg_impedance_magnitude_ohm"],
                    row["reference_impedance_ohm"],
                ]
                for row in ts_summaries
            ],
        )

    if ts_details:
        ws = wb.create_sheet("touchstone_detail")
        write_sheet(
            ws,
            [
                "touchstone_file",
                "port",
                "freq_GHz",
                "vswr",
                "impedance_real_ohm",
                "impedance_imag_ohm",
                "impedance_magnitude_ohm",
                "gamma_real",
                "gamma_imag",
                "reference_impedance_ohm",
            ],
            [
                [
                    row["touchstone_file"],
                    row["port"],
                    maybe_float(float(row["freq_GHz"]), 6),
                    maybe_float(float(row["vswr"]), 6),
                    maybe_float(float(row["impedance_real_ohm"]), 6),
                    maybe_float(float(row["impedance_imag_ohm"]), 6),
                    maybe_float(float(row["impedance_magnitude_ohm"]), 6),
                    maybe_float(float(row["gamma_real"]), 6),
                    maybe_float(float(row["gamma_imag"]), 6),
                    maybe_float(float(row["reference_impedance_ohm"]), 4),
                ]
                for row in ts_details
            ],
        )

    for ws in wb.worksheets:
        autosize_columns(ws)

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)


def main() -> int:
    parser = argparse.ArgumentParser(description="Create an XLSX workbook with extracted FFS and Touchstone metrics.")
    parser.add_argument("output", type=Path, help="Output workbook path (.xlsx).")
    parser.add_argument("ffs", nargs="*", type=Path, help="Optional CST Farfield Source files (.ffs).")
    parser.add_argument("--beam-workbook", type=Path, default=None, help="Optional beam workbook to reuse instead of reparsing .ffs files.")
    parser.add_argument("--touchstone", type=Path, default=None, help="Optional Touchstone file (.s1p/.s2p).")
    parser.add_argument("--smooth", type=int, default=5, help="Smoothing window used for beam efficiency extraction.")
    parser.add_argument("--theta-window", type=float, default=8.0, help="Theta window in degrees used for main-lobe detection.")
    parser.add_argument("--ffs-fmin", type=float, default=None, help="Lower FFS frequency bound in GHz.")
    parser.add_argument("--ffs-fmax", type=float, default=None, help="Upper FFS frequency bound in GHz.")
    parser.add_argument("--touchstone-fmin", type=float, default=None, help="Lower Touchstone frequency bound in GHz.")
    parser.add_argument("--touchstone-fmax", type=float, default=None, help="Upper Touchstone frequency bound in GHz.")
    args = parser.parse_args()

    if not args.ffs and args.touchstone is None:
        raise SystemExit("Provide at least one .ffs file or a --touchstone file.")

    print("Starting extracted-data workbook generation...")
    overview_rows: list[list[object]] = [
        ["output", str(args.output)],
        ["beam_smooth", args.smooth],
        ["theta_window_deg", args.theta_window],
        ["ffs_fmin_GHz", args.ffs_fmin],
        ["ffs_fmax_GHz", args.ffs_fmax],
        ["touchstone_fmin_GHz", args.touchstone_fmin],
        ["touchstone_fmax_GHz", args.touchstone_fmax],
        ["ffs_count", len(args.ffs)],
        ["touchstone_file", str(args.touchstone) if args.touchstone else ""],
    ]

    ffs_summaries: list[dict[str, object]] = []
    ffs_details: list[dict[str, object]] = []
    use_beam_workbook = bool(args.beam_workbook and args.ffs and beam_workbook_is_fresh(args.beam_workbook, args.ffs))
    if use_beam_workbook:
        print(f"Reusing FFS metrics from beam workbook: {args.beam_workbook.name}")
    for ffs_path in args.ffs:
        print(f"Processing FFS: {ffs_path.name}")
        all_rows = compute_ffs_rows_from_beam_workbook(args.beam_workbook, ffs_path) if use_beam_workbook else compute_ffs_rows(ffs_path, args.smooth, args.theta_window)
        selected_rows, used_fmin, used_fmax = filter_rows_by_range(all_rows, args.ffs_fmin, args.ffs_fmax)
        if selected_rows:
            summary = summarize_ffs_rows(selected_rows)
            if summary is not None:
                summary["freq_min_GHz"] = maybe_float(used_fmin, 6)
                summary["freq_max_GHz"] = maybe_float(used_fmax, 6)
                ffs_summaries.append(summary)
            ffs_details.extend(selected_rows)

    ts_summaries: list[dict[str, object]] = []
    ts_details: list[dict[str, object]] = []
    if args.touchstone:
        print(f"Processing Touchstone: {args.touchstone.name}")
        all_rows = compute_touchstone_rows(args.touchstone)
        for port in sorted({str(row["port"]) for row in all_rows}):
            port_rows = [row for row in all_rows if str(row["port"]) == port]
            selected_rows, used_fmin, used_fmax = filter_rows_by_range(port_rows, args.touchstone_fmin, args.touchstone_fmax)
            if not selected_rows:
                continue
            summary = summarize_touchstone_rows(selected_rows)
            if summary is not None:
                summary["freq_min_GHz"] = maybe_float(used_fmin, 6)
                summary["freq_max_GHz"] = maybe_float(used_fmax, 6)
                ts_summaries.append(summary)
            ts_details.extend(selected_rows)

    if not ffs_summaries and not ts_summaries:
        raise SystemExit("No usable data could be extracted from the provided inputs.")

    build_workbook(args.output, overview_rows, ffs_summaries, ffs_details, ts_summaries, ts_details)
    print(f"Wrote extracted workbook: {args.output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
